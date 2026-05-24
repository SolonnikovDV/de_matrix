# -*- coding: utf-8 -*-
"""
Единый источник данных: JSON, YAML, Excel.
Структура матрицы — только дерево ``nodes``; action_templates, literature, examples, ui_config — как раньше.
Стили (name, icon, color) и инструменты (tools) не в источнике: загружаются из config/metadata.yaml,
инструменты привязываются к листьям по совпадению паттернов в тексте.
Валидация структуры — core/schema.py.
"""
import json
import math
import csv
from copy import deepcopy
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

from .tree import strip_transient_node_fields, tabular_hierarchy_to_nodes
from .schema import validate_source, ValidationResult, SCHEMA_VERSION
from .matrix_schema import normalize_level_tags, normalize_responsible_value
from .excel_unified_relational import try_load_unified_relational_xlsx
from .tabular_matrix_contract import (
    exls_tabular_rows_to_nodes,
    matches_exls_tabular_columns,
)

# Ключи в источнике (без stack_labels и action_tools — они в config/metadata.yaml).


def excel_cell_str(value: Any) -> str:
    """
    Безопасная строка из ячейки Excel.
    pandas/openpyxl отдают int/float; выражение (v or '').strip() падает на truthy float (например 3.0).
    """
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    return s
META_KEYS = (
    "action_examples",
    "literature",
    "action_templates",
    "ui_config",
)


def _parse_level_tag(value: Any) -> Optional[str]:
    """Нормализует уровень компетенции."""
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    aliases = {
        "junior": "junior",
        "jr": "junior",
        "middle": "middle",
        "mid": "middle",
        "senior": "senior",
        "sr": "senior",
        "джуниор": "junior",
        "мидл": "middle",
        "сеньор": "senior",
    }
    return aliases.get(text, text)


def _parse_review_questions(value: Any) -> List[str]:
    """Парсит строку/массив проверочных вопросов в список строк."""
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    text = str(value).strip()
    if not text:
        return []
    text = text.replace("\r\n", "\n")
    if ";" in text:
        parts = [p.strip() for p in text.split(";")]
    else:
        parts = [p.strip() for p in text.split("\n")]
    return [p for p in parts if p]


def load_json(path: str) -> Dict:
    """Загружает JSON-файл."""
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_yaml(path: str) -> Dict:
    """Загружает YAML-файл."""
    try:
        import yaml
    except ImportError:
        raise RuntimeError("Для загрузки YAML установите: pip install pyyaml")
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _col(columns: List[str], *candidates: str) -> Optional[str]:
    for c in candidates:
        if c in columns:
            return c
    return None


def _merge_multiline_text(existing: str, incoming: str) -> str:
    left = str(existing or "").strip()
    right = str(incoming or "").strip()
    if not right:
        return left
    if not left:
        return right
    if right in left:
        return left
    return f"{left}\n{right}"


def _normalize_tabular_column_name(name: str) -> str:
    return str(name or "").strip().casefold()


# Колонки контента листа (не уровни дерева). Порядок колонок слева направо: item → item → … → leaf_view.
_TABULAR_METADATA_COLUMN_NAMES = frozenset(
    _normalize_tabular_column_name(x)
    for x in (
        "Status",
        "Статус",
        "Questions",
        "Вопросы",
        "Review Questions",
        "Проверочные вопросы",
        "Reviewer Questions",
        "Вопросы ревьюера",
        "reviewer_questions",
        "Materials",
        "Материалы",
        "Tasks",
        "Задачи",
        "Optional",
        "Опционально",
        "Опционально для уровня",
        "Author",
        "Автор",
        "Reviewer",
        "Ревьюер",
        "Responsible",
        "Ответственный",
        "responsible",
        "Description",
        "description",
        "Описание",
        "Описание навыка",
        "Template ID",
        "template_id",
        "Template_ID",
        "Level Tag",
        "level_tag",
        "Level",
        "Уровень",
        "Level Tags",
        "level_tags",
        "Наклейки уровня",
        "Action Level Tags",
        "action_level_tags",
        "Subaction Level Tags",
        "subaction_level_tags",
        "Skill Sticker",
        "skill_sticker",
        "Level Sticker",
        "level_sticker",
        "Наклейка уровня",
    )
)


def _is_tabular_metadata_column(col: str) -> bool:
    return _normalize_tabular_column_name(col) in _TABULAR_METADATA_COLUMN_NAMES


def _detect_tabular_hierarchy_columns(columns: List[str]) -> List[str]:
    """
    Иерархия item-колонок: слева направо, до первой leaf/metadata-колонки.
    Старший уровень — крайняя слева колонка; вправо — вложенность.
    """
    hierarchy: List[str] = []
    for col in columns:
        name = str(col or "").strip()
        if not name:
            continue
        if _is_tabular_metadata_column(name):
            break
        hierarchy.append(name)
    return hierarchy


def _effective_tabular_hierarchy_path(last_by_col: Dict[str, str], hierarchy_cols: List[str]) -> List[str]:
    """Путь узлов с carry-forward; ведущие пустые уровни отбрасываются."""
    path = [last_by_col.get(col, "") for col in hierarchy_cols]
    while path and not path[0]:
        path.pop(0)
    return path


def _attach_tabular_leaf_fields(
    node: Dict[str, Any],
    row: Dict[str, Any],
    hierarchy_cols: List[str],
    all_columns: List[str],
    *,
    owner_col: Optional[str],
    questions_col: Optional[str],
    reviewer_questions_col: Optional[str],
    level_col: Optional[str],
    level_tags_col: Optional[str],
    sticker_col: Optional[str],
    tpl_col: Optional[str],
    desc_col: Optional[str],
) -> None:
    """Метаданные строки — на лист (leaf_view и поля узла)."""
    control = set(hierarchy_cols)
    if owner_col:
        control.add(owner_col)
    if questions_col:
        control.add(questions_col)
    if reviewer_questions_col:
        control.add(reviewer_questions_col)

    extra_view: Dict[str, str] = {}
    for col_name in all_columns:
        if col_name in control:
            continue
        value = excel_cell_str(row.get(col_name))
        if value:
            extra_view[str(col_name).strip()] = value
    if extra_view:
        lv = node.get("leaf_view")
        if not isinstance(lv, dict):
            lv = {}
        for k, v in extra_view.items():
            prev = lv.get(k)
            if isinstance(prev, str):
                lv[k] = _merge_multiline_text(prev, v)
            elif prev:
                lv[k] = prev
            else:
                lv[k] = v
        node["leaf_view"] = lv

    if desc_col:
        desc = excel_cell_str(row.get(desc_col))
        if desc and not node.get("description"):
            node["description"] = desc
    if owner_col:
        responsible = normalize_responsible_value(excel_cell_str(row.get(owner_col)))
        if responsible and not node.get("responsible"):
            node["responsible"] = responsible
    if sticker_col:
        sticker = excel_cell_str(row.get(sticker_col)).lower()
        if sticker and not node.get("level_sticker"):
            node["level_sticker"] = sticker

    rq: List[str] = []
    if questions_col:
        rq.extend(_parse_review_questions(row.get(questions_col)))
    if reviewer_questions_col:
        for q in _parse_review_questions(row.get(reviewer_questions_col)):
            if q not in rq:
                rq.append(q)
    if rq:
        existing = [str(q).strip() for q in (node.get("review_questions") or []) if str(q).strip()]
        for q in rq:
            if q not in existing:
                existing.append(q)
        node["review_questions"] = existing

    if level_tags_col:
        ltags = normalize_level_tags(row.get(level_tags_col))
    elif level_col:
        ltags = normalize_level_tags(row.get(level_col))
    else:
        ltags = []
    if ltags:
        node["level_tags"] = ltags

    if tpl_col:
        tpl = excel_cell_str(row.get(tpl_col)) or None
        if tpl and not node.get("template_id"):
            node["template_id"] = tpl


def _nested_dict_to_generic_nodes(tree: Dict[str, Dict[str, Any]]) -> List[Dict]:
    out: List[Dict] = []
    for name in sorted(tree.keys(), key=lambda x: tree[x].get("__order__", 0)):
        slot = tree[name]
        children_tree = slot.get("__children__")
        if not isinstance(children_tree, dict):
            children_tree = {}
        children = _nested_dict_to_generic_nodes(children_tree)
        node: Dict[str, Any] = {"name": name, "children": children}
        for key in ("description", "responsible", "level_sticker", "template_id", "level_tags", "review_questions", "leaf_view"):
            if slot.get(key) is not None:
                node[key] = deepcopy(slot[key])
        out.append(node)
    return out


def _tabular_left_to_right_hierarchy_to_nodes(
    rows: List[Dict[str, Any]],
    columns: List[str],
    hierarchy_cols: List[str],
) -> List[Dict]:
    """
    Табличная иерархия exls_matrix: колонки item слева направо, null = наследование сверху.
    """
    if not hierarchy_cols:
        return []

    owner_col = _col(columns, "Responsible", "responsible", "Ответственный", "Автор")
    desc_col = _col(columns, "Description", "description", "Описание навыка", "Описание")
    sticker_col = _col(columns, "Skill Sticker", "skill_sticker", "Level Sticker", "level_sticker", "Наклейка уровня")
    tpl_col = _col(columns, "Template ID", "template_id", "Template_ID")
    level_col = _col(columns, "Level Tag", "level_tag", "Level", "Уровень")
    level_tags_col = _col(columns, "Level Tags", "level_tags", "Наклейки уровня")
    questions_col = _col(columns, "Review Questions", "review_questions", "Проверочные вопросы", "Вопросы")
    reviewer_questions_col = _col(columns, "Вопросы ревьюера", "Reviewer Questions", "reviewer_questions")

    last_by_col: Dict[str, str] = {col: "" for col in hierarchy_cols}
    root: Dict[str, Dict[str, Any]] = {}
    order_counter = 0

    for row in rows:
        for col in hierarchy_cols:
            val = excel_cell_str(row.get(col))
            if val:
                last_by_col[col] = val

        path = _effective_tabular_hierarchy_path(last_by_col, hierarchy_cols)
        if not path or not path[-1]:
            continue

        level = root
        for depth, segment in enumerate(path):
            if segment not in level:
                level[segment] = {"__order__": order_counter, "__children__": {}}
                order_counter += 1
            slot = level[segment]
            if depth == len(path) - 1:
                _attach_tabular_leaf_fields(
                    slot,
                    row,
                    hierarchy_cols,
                    columns,
                    owner_col=owner_col,
                    questions_col=questions_col,
                    reviewer_questions_col=reviewer_questions_col,
                    level_col=level_col,
                    level_tags_col=level_tags_col,
                    sticker_col=sticker_col,
                    tpl_col=tpl_col,
                    desc_col=desc_col,
                )
            child_level = slot.get("__children__")
            if not isinstance(child_level, dict):
                child_level = {}
                slot["__children__"] = child_level
            level = child_level

    return strip_transient_node_fields(_nested_dict_to_generic_nodes(root))


def _tabular_rows_to_nodes(rows: List[Dict[str, Any]], columns: List[str]) -> Dict:
    """Общий парсер табличных строк (CSV / XLSX / JSON table) → nodes + ui_config."""
    if matches_exls_tabular_columns(columns):
        nodes, ui_config = exls_tabular_rows_to_nodes(
            rows,
            columns,
            cell_str=excel_cell_str,
            strip_transient=strip_transient_node_fields,
        )
        return {"nodes": nodes, "ui_config": ui_config}

    hierarchy_cols = _detect_tabular_hierarchy_columns(columns)
    if hierarchy_cols:
        return {"nodes": _tabular_left_to_right_hierarchy_to_nodes(rows, columns, hierarchy_cols), "ui_config": {}}

    return {"nodes": [], "ui_config": {}}


def load_excel(path: str, sheet_name: Optional[str] = None) -> Dict:
    """
    Загружает Excel-файл. Ожидаемые колонки: Domain, Skill, Action, [Subaction], [Description], [Template ID], ...
    Или первый столбец — уровень вложенности (1=домен, 2=навык, 3=действие, 4=поддействие), далее Name, Description, Template ID.
    """
    rows: List[Dict[str, str]] = []
    columns: List[str] = []

    # 1) Быстрый путь через pandas (если установлен)
    try:
        import pandas as pd
        df = pd.read_excel(path, sheet_name=sheet_name or 0)
        df = df.astype(str).replace("nan", "")
        columns = [str(c) for c in df.columns]
        rows = []
        for _, row in df.iterrows():
            rows.append({str(k): excel_cell_str(row.get(k)) for k in df.columns})
    except ImportError:
        # 2) Фолбэк без pandas — только openpyxl
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("Для загрузки Excel установите: pip install openpyxl")

        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            try:
                if isinstance(sheet_name, str):
                    ws = wb[sheet_name]
                elif isinstance(sheet_name, int):
                    ws = wb.worksheets[sheet_name]
                else:
                    ws = wb.worksheets[0]
            except Exception:
                ws = wb.worksheets[0]

            header_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
            header_row = next(header_iter, None)
            if not header_row:
                return {"nodes": []}

            columns = [str(c).strip() if c is not None else "" for c in header_row]
            rows = []
            max_data_rows = 200000
            empty_streak = 0
            max_empty_streak = 2000
            for raw in ws.iter_rows(min_row=2, values_only=True):
                rec: Dict[str, str] = {}
                has_values = False
                for i, col_name in enumerate(columns):
                    if not col_name:
                        continue
                    v = raw[i] if i < len(raw) else None
                    s = excel_cell_str(v)
                    if s:
                        has_values = True
                    rec[col_name] = s

                if has_values:
                    rows.append(rec)
                    empty_streak = 0
                else:
                    empty_streak += 1
                    # У части файлов XLSX "хвост" состоит из тысяч пустых строк из-за форматирования листа.
                    # Прерываем чтение, чтобы preview/validation не зависали.
                    if empty_streak >= max_empty_streak:
                        break

                if len(rows) >= max_data_rows:
                    break
        finally:
            wb.close()

    return _tabular_rows_to_nodes(rows, columns)


def load_csv(path: str) -> Dict:
    """Загружает CSV-файл (поддержка ; , tab, utf-8/utf-8-sig/cp1251)."""
    encodings = ("utf-8-sig", "utf-8", "cp1251")
    last_error: Optional[Exception] = None
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc, newline="") as fp:
                sample = fp.read(8192)
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
        except Exception as exc:
            last_error = exc
            continue

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ";" if sample.count(";") >= sample.count(",") else ","

        with open(path, "r", encoding=enc, newline="") as fp:
            reader = csv.DictReader(fp, delimiter=delimiter)
            columns = [str(c).strip() for c in (reader.fieldnames or []) if str(c).strip()]
            rows: List[Dict[str, str]] = []
            for row in reader:
                if not isinstance(row, dict):
                    continue
                rec: Dict[str, str] = {}
                has_values = False
                for col_name in columns:
                    val = excel_cell_str(row.get(col_name))
                    if val:
                        has_values = True
                    rec[col_name] = val
                if has_values:
                    rows.append(rec)
            return _tabular_rows_to_nodes(rows, columns)

    if last_error:
        raise RuntimeError(f"Ошибка чтения CSV: {last_error}")
    return {"nodes": [], "ui_config": {}}


def _level_outline_to_domain_rows(children: List[Dict]) -> List[Dict]:
    """Иерархия по колонке Level → промежуточные строки домен/навык/действие (только для Excel)."""
    domains = []
    for d in children:
        domain = {"name": d.get("name", ""), "skills": []}
        for s in d.get("children", []):
            skill = {
                "name": s.get("name", ""),
                "description": s.get("description", ""),
                "responsible": s.get("responsible", ""),
                "level_sticker": s.get("level_sticker", ""),
                "actions": [],
            }
            for a in s.get("children", []):
                if a.get("children"):
                    action = {"text": a.get("name", ""), "template_id": a.get("template_id"), "subactions": []}
                    if a.get("level_tag"):
                        action["level_tag"] = a.get("level_tag")
                    if a.get("review_questions"):
                        action["review_questions"] = a.get("review_questions")
                    for sub in a["children"]:
                        sub_item = {
                            "text": sub.get("name", ""),
                            "template_id": sub.get("template_id"),
                        }
                        if sub.get("level_tag"):
                            sub_item["level_tag"] = sub.get("level_tag")
                        if sub.get("review_questions"):
                            sub_item["review_questions"] = sub.get("review_questions")
                        action["subactions"].append(sub_item)
                    skill["actions"].append(action)
                else:
                    action_item = {"text": a.get("name", ""), "template_id": a.get("template_id")}
                    if a.get("level_tag"):
                        action_item["level_tag"] = a.get("level_tag")
                    if a.get("review_questions"):
                        action_item["review_questions"] = a.get("review_questions")
                    skill["actions"].append(action_item)
            domain["skills"].append(skill)
        domains.append(domain)
    return domains


def _normalize_domain_rows_for_sheet(domains_list: List[Dict]) -> List[Dict]:
    """Нормализует промежуточные domain/skill/action-строки из классического Excel."""
    normalized_domains: List[Dict] = []
    for d in domains_list:
        if not isinstance(d, dict):
            continue
        domain = {k: v for k, v in d.items() if k != "skills"}
        domain["skills"] = []
        for s in d.get("skills", []):
            if not isinstance(s, dict):
                continue
            skill = {k: v for k, v in s.items() if k != "actions"}
            skill["actions"] = [_normalize_action(a) for a in s.get("actions", [])]
            domain["skills"].append(skill)
        normalized_domains.append(domain)
    _sanitize_responsible_fields_in_domains(normalized_domains)
    return normalized_domains


def _excel_domain_rows_as_nodes(domains_list: List[Dict]) -> List[Dict]:
    """Классический Excel → generic ``nodes`` (без id/path/is_leaf)."""
    norm = _normalize_domain_rows_for_sheet(domains_list)
    raw = tabular_hierarchy_to_nodes(norm)
    return strip_transient_node_fields(raw)


def _empty_unified() -> Dict:
    """Пустой единый источник (только структура и текстовое содержимое)."""
    return {
        "schema_version": SCHEMA_VERSION,
        "domains": [],
        "nodes": [],
        "action_examples": [],
        "literature": {},
        "action_templates": {},
        "ui_config": {},
    }


def _normalize_generic_node(n: Dict) -> Dict:
    """Узел дерева произвольной глубины для unified (имена из файла / JSON)."""
    if not isinstance(n, dict):
        return {"name": "", "children": []}
    name = (n.get("name") or n.get("text") or "").strip()
    out: Dict = {"name": name, "children": []}
    for key in ("description", "responsible", "level_sticker", "code", "template_id", "excel_path_key"):
        if key in n and n[key]:
            if key == "excel_path_key":
                out[key] = str(n[key]).strip()
            elif key == "level_sticker":
                out[key] = str(n[key]).strip().lower()
            else:
                out[key] = str(n[key]).strip() if isinstance(n[key], str) else n[key]
    ltags = normalize_level_tags(n.get("level_tags") or n.get("level_tag"))
    if ltags:
        out["level_tags"] = ltags
        if len(ltags) == 1:
            out["level_tag"] = ltags[0]
    elif n.get("level_tag"):
        out["level_tag"] = _parse_level_tag(n.get("level_tag"))
    for key in ("section", "status", "author", "reviewer", "skill_sections"):
        if key in n and n[key]:
            if key == "skill_sections":
                out[key] = deepcopy(n[key])
            elif isinstance(n[key], str):
                out[key] = n[key].strip()
            else:
                out[key] = n[key]
    ch = n.get("children") or []
    if ch:
        out["children"] = [_normalize_generic_node(x) for x in ch if isinstance(x, dict)]
    lv = n.get("leaf_view")
    if isinstance(lv, dict) and lv:
        out["leaf_view"] = deepcopy(lv)
    rq = _parse_review_questions(n.get("review_questions"))
    if rq:
        out["review_questions"] = rq
    return out


def _sanitize_responsible_fields_in_domains(domains: List[Dict]) -> None:
    """Убирает заглушки вроде «не указан» из полей responsible (импорт / нормализация)."""
    for d in domains:
        if not isinstance(d, dict):
            continue
        if "responsible" in d:
            d["responsible"] = normalize_responsible_value(d.get("responsible"))
        for s in d.get("skills") or []:
            if not isinstance(s, dict):
                continue
            if "responsible" in s:
                s["responsible"] = normalize_responsible_value(s.get("responsible"))
            for act in s.get("actions") or []:
                if not isinstance(act, dict):
                    continue
                if "responsible" in act:
                    act["responsible"] = normalize_responsible_value(act.get("responsible"))
                for sub in act.get("subactions") or []:
                    if isinstance(sub, dict) and "responsible" in sub:
                        sub["responsible"] = normalize_responsible_value(sub.get("responsible"))


def _normalize_action(a: Dict) -> Dict:
    """Приводит действие к стандартному формату {text, template_id?, subactions?}."""
    if a.get("type") == "group":
        return {
            "text": a.get("name", ""),
            "template_id": None,
            "subactions": [_normalize_subaction(s) for s in (a.get("items") or [])],
        }
    subactions = a.get("subactions")
    ltags = normalize_level_tags(a.get("level_tags") or a.get("level_tag"))
    normalized_questions = _parse_review_questions(a.get("review_questions"))
    if subactions:
        out = {**a, "subactions": [_normalize_subaction(s) for s in subactions]}
    else:
        out = dict(a)
    if ltags:
        out["level_tags"] = ltags
        if len(ltags) == 1:
            out["level_tag"] = ltags[0]
        else:
            out.pop("level_tag", None)
    else:
        out.pop("level_tags", None)
    if normalized_questions:
        out["review_questions"] = normalized_questions
    elif "review_questions" in out and not out["review_questions"]:
        out.pop("review_questions", None)
    lv = a.get("leaf_view")
    if isinstance(lv, dict) and lv:
        out["leaf_view"] = lv
    return out


def _normalize_subaction(s: Dict) -> Dict:
    """Приводит поддействие к формату {text, template_id?, level_tag?, review_questions?}."""
    if isinstance(s, dict):
        out = {"text": s.get("text", s.get("name", "")), "template_id": s.get("template_id")}
        stags = normalize_level_tags(s.get("level_tags") or s.get("level_tag"))
        if stags:
            out["level_tags"] = stags
            if len(stags) == 1:
                out["level_tag"] = stags[0]
        elif s.get("level_tag"):
            out["level_tag"] = _parse_level_tag(s.get("level_tag"))
        questions = _parse_review_questions(s.get("review_questions"))
        if questions:
            out["review_questions"] = questions
        lv = s.get("leaf_view")
        if isinstance(lv, dict) and lv:
            out["leaf_view"] = lv
        return out
    return {"text": str(s), "template_id": None}


def _merge_tabular_ui_config(out: Dict, tabular: Dict) -> None:
    ui = tabular.get("ui_config")
    if isinstance(ui, dict) and ui:
        out.setdefault("ui_config", {})
        if isinstance(out["ui_config"], dict):
            for k, v in ui.items():
                if v is not None:
                    out["ui_config"][k] = deepcopy(v)


def _normalize_unified(data: Any) -> Dict:
    """Приводит загруженные данные к формату единого источника (без стилей и списков инструментов)."""
    # Поддержка tabular JSON list-of-records.
    if isinstance(data, list):
        rows = [x for x in data if isinstance(x, dict)]
        columns: List[str] = []
        seen = set()
        for row in rows:
            for key in row.keys():
                k = str(key).strip()
                if not k or k in seen:
                    continue
                seen.add(k)
                columns.append(k)
        out = _empty_unified()
        out["schema_version"] = SCHEMA_VERSION
        tabular = _tabular_rows_to_nodes(rows, columns)
        out["nodes"] = tabular.get("nodes") or []
        _merge_tabular_ui_config(out, tabular)
        return out
    if not isinstance(data, dict):
        return _empty_unified()
    out = _empty_unified()
    out["schema_version"] = data.get("schema_version", SCHEMA_VERSION)
    nodes_in = data.get("nodes")
    if isinstance(nodes_in, list) and nodes_in:
        out["nodes"] = [_normalize_generic_node(x) for x in nodes_in if isinstance(x, dict)]
    else:
        domains_in = data.get("domains")
        if isinstance(domains_in, list) and domains_in:
            out["nodes"] = _excel_domain_rows_as_nodes(domains_in)
        else:
            # Поддержка JSON workbook dump: {"source_file": "...", "sheets": {name: {columns, rows}}}
            sheets = data.get("sheets")
            if isinstance(sheets, dict) and sheets:
                first_sheet = next((v for v in sheets.values() if isinstance(v, dict)), None)
                rows = first_sheet.get("rows") if isinstance(first_sheet, dict) else None
                columns = first_sheet.get("columns") if isinstance(first_sheet, dict) else None
                if isinstance(rows, list):
                    rows_norm = [x for x in rows if isinstance(x, dict)]
                    if not isinstance(columns, list) or not columns:
                        col_seen = set()
                        col_list: List[str] = []
                        for row in rows_norm:
                            for key in row.keys():
                                k = str(key).strip()
                                if not k or k in col_seen:
                                    continue
                                col_seen.add(k)
                                col_list.append(k)
                        columns = col_list
                    cols_clean = [str(c).strip() for c in columns if str(c).strip()]
                    tabular = _tabular_rows_to_nodes(rows_norm, cols_clean)
                    out["nodes"] = tabular.get("nodes") or []
                    _merge_tabular_ui_config(out, tabular)
                else:
                    out["nodes"] = []
            else:
                out["nodes"] = []
    out["domains"] = []
    for key in META_KEYS:
        if key in data and data[key] is not None:
            default = [] if key == "action_examples" else {}
            out[key] = data[key] if isinstance(data[key], (dict, list)) else default
    return out


def load_unified_source(
    source_path: str,
    source_type: Optional[str] = None,
    validate: bool = False,
) -> Dict:
    """
    Загружает единый источник данных (JSON, YAML или Excel).
    Структура — ``nodes``; метаданные — шаблоны, литература, стек, ui_config.
    Для Excel загружается только структура; мета при необходимости — из того же файла (листы) или пустая.
    Если validate=True, при ошибках валидации выбрасывается ValueError с текстом ошибок.
    """
    path = Path(source_path)
    if not path.exists():
        return _empty_unified()

    ext = (source_type or path.suffix or "").lower()
    if ext == ".json":
        data = load_json(str(path))
    elif ext in (".yaml", ".yml"):
        data = load_yaml(str(path))
    elif ext == ".csv":
        data = _load_csv_as_unified(str(path))
    elif ext in (".xlsx", ".xls"):
        data = _load_excel_as_unified(str(path))
    else:
        data = load_json(str(path))

    out = _normalize_unified(data)
    if validate:
        vr = validate_source(out)
        if not vr.ok:
            raise ValueError("Ошибки валидации: " + "; ".join(vr.errors))
    return out


def load_unified_source_with_validation(
    source_path: str, source_type: Optional[str] = None
) -> Tuple[Dict, ValidationResult]:
    """
    Загружает источник и возвращает (data, validation_result).
    Не выбрасывает исключение при ошибках валидации.
    """
    path = Path(source_path)
    if not path.exists():
        return _empty_unified(), ValidationResult(ok=False, errors=["Файл не найден"])

    try:
        ext = (source_type or path.suffix or "").lower()
        if ext == ".json":
            data = load_json(str(path))
        elif ext in (".yaml", ".yml"):
            data = load_yaml(str(path))
        elif ext == ".csv":
            data = _load_csv_as_unified(str(path))
        elif ext in (".xlsx", ".xls"):
            data = _load_excel_as_unified(str(path))
        else:
            data = load_json(str(path))
    except Exception as e:
        return _empty_unified(), ValidationResult(ok=False, errors=[str(e)])

    out = _normalize_unified(data)
    vr = validate_source(out)
    return out, vr


def _load_csv_as_unified(path: str) -> Dict:
    structure = load_csv(path)
    out: Dict[str, Any] = {"nodes": structure.get("nodes", [])}
    ui = structure.get("ui_config")
    if isinstance(ui, dict) and ui:
        out["ui_config"] = deepcopy(ui)
    return out


def _load_excel_as_unified(path: str) -> Dict:
    """Читает Excel: unified relational (теги в шапке) или табличный exls (Domain/Skill/…)."""
    ur = try_load_unified_relational_xlsx(path)
    if ur:
        return {
            "nodes": ur.get("nodes") or [],
            "ui_config": ur.get("ui_config") or {},
        }
    structure = load_excel(path)
    out: Dict[str, Any] = {"nodes": structure.get("nodes", [])}
    ui = structure.get("ui_config")
    if isinstance(ui, dict) and ui:
        out["ui_config"] = deepcopy(ui)
    return out


def load_excel_for_matrix_import(path: str) -> Dict:
    """Предпросмотр/импорт Excel: сначала unified relational (enriched), иначе классический лист."""
    return _load_excel_as_unified(path)


def load_csv_for_matrix_import(path: str) -> Dict:
    """Предпросмотр/импорт CSV: табличный формат -> unified nodes."""
    return _load_csv_as_unified(path)


def load_matrix(source_path: str, source_type: Optional[str] = None) -> Dict:
    """
    Загружает матрицу из файла (только структура). Для единого источника используйте load_unified_source.
    """
    unified = load_unified_source(source_path, source_type)
    return {"nodes": unified.get("nodes") or []}
