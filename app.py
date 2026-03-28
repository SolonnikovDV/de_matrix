import json
import csv
import io
import html as html_lib
import os
import re
import hashlib
import sys
import argparse
import socket
import smtplib
import secrets
from datetime import datetime, timezone, timedelta
from email.message import EmailMessage
from flask import Flask, render_template, jsonify, abort, send_from_directory, request, session, redirect, url_for, Response
from werkzeug.security import generate_password_hash, check_password_hash

from copy import deepcopy
from pathlib import Path as PathLib
from typing import Dict, Optional, Tuple, Any, List, Set

from core.tree import (
    assign_paths_to_generic_nodes,
    build_tree_from_matrix_data,
    collect_leaves,
    get_node_by_path,
    get_ancestors,
    path_to_url,
    strip_transient_node_fields,
)
from core.loaders import load_unified_source, load_excel_for_matrix_import, META_KEYS
from core.schema import validate_source, get_schema_info
from core.matrix_schema import (
    action_level_tags_for_json,
    build_constructor_levels,
    effective_matrix_column_schema,
    matrix_roundtrip_header_cell,
    merge_matrix_levels,
    normalize_level_tags,
    schema_entries_for_ui,
    subaction_level_tags_for_json,
    STICKER_GRADES,
    TAG_SKILL_STICKER,
)
from core import config_loader as _config_loader
from core.config_loader import load_app_config, load_metadata, invalidate_metadata_cache
from core.tools_matcher import get_tools_for_text
from core.checkpoint import (
    load_checkpoint,
    save_checkpoint,
    source_matches_checkpoint,
    build_checkpoint_data,
    get_source_content_hash,
)
from core.backup import (
    create_backup,
    list_backups,
    restore_backup,
    check_backup_compatibility,
    get_stable_backup_id,
    set_stable_backup_id,
    ensure_stable_backup,
)
from core.upload_merge import merge_upload_into_source
from core.excel_unified_export import build_unified_export_table
from core.diff_engine import build_revision_payload
from storage.runtime import get_storage_mode, approval_required as storage_approval_required
from storage.db import db_session, ENGINE
from storage.models import (
    Base,
    MATRIX_STRUCT_SCHEMA,
    ChangeRequest,
    ChangeRevision,
    ApprovalDecision,
    User,
    ChangeDiscussionThread,
    ChangeDiscussionMessage,
    NotificationLog,
    UserPresenceSession,
)
from sqlalchemy import select, text, func, or_, asc, desc
from storage.approval_repo import (
    create_change_request,
    add_revision,
    set_status as approval_set_status,
    get_latest_payload,
    list_change_requests,
    get_change_request_details,
    leaf_path_hints_from_applied_changes,
)
from storage.postgres_repo import (
    load_unified_from_db,
    replace_unified_in_db,
    load_matrix_nodes_nested,
)
from storage.staging_repo import create_staging_batch, load_staging_tree_projection
from storage.mongo_repo import (
    load_literature_map,
    upsert_literature_item,
    delete_literature_item as mongo_delete_literature_item,
    add_literature_file as mongo_add_literature_file,
)

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
AUTHOR_NAME = os.environ.get("DE_MATRIX_AUTHOR_NAME", "Dmitry Solonnikov")
AUTHOR_TELEGRAM = os.environ.get("DE_MATRIX_AUTHOR_TELEGRAM", "https://t.me/Dmitry_as_SoloD")
REPOSITORY_URL = os.environ.get("DE_MATRIX_REPO_URL", "https://github.com/SolonnikovDV/de_matrix.git")
SECRET_KEY = os.environ.get("DE_MATRIX_SECRET_KEY") or hashlib.sha256(os.urandom(32)).hexdigest()
ADMIN_USERNAME = os.environ.get("DE_MATRIX_ADMIN_USERNAME", "admin")
ADMIN_BOOTSTRAP_PASSWORD = os.environ.get("DE_MATRIX_ADMIN_PASSWORD", "")
ADMIN_DISPLAY_NAME = (os.environ.get("DE_MATRIX_ADMIN_FULL_NAME") or "System Administrator").strip() or "System Administrator"
SYSTEM_ADMIN_NAME = "System Administrator"
APP_USER_NAME = "App User"
CUSTOM_ADMIN_NAME = "Custom Administrator"
E2E_ADMIN_USERNAME = "e2e_admin"
TRUST_REQUEST_ROLE = os.environ.get("DE_MATRIX_TRUST_REQUEST_ROLE", "0").strip().lower() in ("1", "true", "yes")
AUTH_REQUIRED = os.environ.get("DE_MATRIX_AUTH_REQUIRED", "1").strip().lower() in ("1", "true", "yes")
NOTIFICATIONS_ENABLED = os.environ.get("DE_MATRIX_NOTIFICATIONS_ENABLED", "1").strip().lower() in ("1", "true", "yes")
SMTP_HOST = (os.environ.get("DE_MATRIX_SMTP_HOST") or "smtp").strip()
SMTP_PORT = int((os.environ.get("DE_MATRIX_SMTP_PORT") or "1025").strip())
SMTP_FROM = (os.environ.get("DE_MATRIX_SMTP_FROM") or "de-matrix@localhost").strip()
PRESENCE_ONLINE_SECONDS = int((os.environ.get("DE_MATRIX_PRESENCE_ONLINE_SECONDS") or "120").strip())
PRESENCE_AWAY_SECONDS = int((os.environ.get("DE_MATRIX_PRESENCE_AWAY_SECONDS") or "900").strip())
app.config["SECRET_KEY"] = SECRET_KEY
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
if os.environ.get("DE_MATRIX_DEPLOY_TARGET", "local").strip().lower() == "production":
    app.config["SESSION_COOKIE_SECURE"] = True
# Путь к config — гарантированно относительно app.py (для metadata.yaml: stack_labels, tools)
_config_loader.CONFIG_DIR = PathLib(BASE_DIR) / "config"

_matrix = None
_meta = None
_tree = None
_current_source_file = None  # имя файла-источника, по которому загружены данные
_admin_seed_checked = False


def _is_db_mode() -> bool:
    return get_storage_mode() == "db"


def _session_actor_role() -> Tuple[Optional[str], Optional[str]]:
    actor = (session.get("actor") or "").strip() or None
    role = (session.get("role") or "").strip().lower() or None
    if role not in ("user", "admin"):
        role = None
    return actor, role


def _load_user(username: str) -> Optional[User]:
    uname = (username or "").strip()
    if not uname:
        return None
    with db_session() as db:
        return db.execute(select(User).where(User.username == uname)).scalars().first()


def _require_authenticated():
    actor = (session.get("actor") or "").strip()
    role = (session.get("role") or "").strip().lower()
    if not session.get("authenticated") or role not in ("user", "admin") or not actor:
        return None, (jsonify({"ok": False, "error": "Authentication required"}), 401)
    return {"actor": actor, "role": role}, None


def _ensure_admin_seed():
    with db_session() as db:
        admin = db.execute(select(User).where(User.username == ADMIN_USERNAME)).scalars().first()
        if admin:
            if (admin.full_name or "").strip() != ADMIN_DISPLAY_NAME:
                admin.full_name = ADMIN_DISPLAY_NAME
                admin.updated_at = _utcnow()
            return
        password = ADMIN_BOOTSTRAP_PASSWORD.strip() or "admin12345"
        db.add(
            User(
                username=ADMIN_USERNAME,
                role="admin",
                full_name=ADMIN_DISPLAY_NAME,
                email="",
                password_hash=generate_password_hash(password),
                must_change_password=True,
                is_active=True,
            )
        )
        print(f"[auth] bootstrap admin '{ADMIN_USERNAME}' created (must change password on first login)")


def _is_safe_next_url(next_url: str) -> bool:
    return bool(next_url) and next_url.startswith("/") and not next_url.startswith("//")


def _extract_actor_role(data: Optional[Dict] = None) -> Tuple[str, str]:
    session_actor, session_role = _session_actor_role()
    if session_role:
        return session_actor or "user", session_role

    payload = data or {}
    actor = (
        request.headers.get("X-Actor")
        or payload.get("actor")
        or request.form.get("actor")
        or request.args.get("actor")
        or "user"
    )
    if TRUST_REQUEST_ROLE:
        role = (
            request.headers.get("X-Role")
            or payload.get("role")
            or request.form.get("role")
            or request.args.get("role")
            or "user"
        )
    else:
        role = "user"
    actor = (actor or "user").strip() or "user"
    role = (role or "user").strip().lower() or "user"
    return actor, role


def _require_admin(data: Optional[Dict] = None):
    actor, role = _extract_actor_role(data)
    if role != "admin":
        return None, (jsonify({"ok": False, "error": "Admin role required"}), 403)
    return actor, None


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _is_valid_email_mask(email: str) -> bool:
    value = (email or "").strip()
    if not value:
        return True
    return bool(re.match(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$", value))


def _default_full_name_for_role(role: str) -> str:
    return CUSTOM_ADMIN_NAME if (role or "").strip().lower() == "admin" else APP_USER_NAME


def _is_system_admin_username(username: str) -> bool:
    return (username or "").strip() == (ADMIN_USERNAME or "").strip()


def _is_e2e_admin_username(username: str) -> bool:
    return (username or "").strip().lower() == E2E_ADMIN_USERNAME


def _client_ip() -> str:
    forwarded = (request.headers.get("X-Forwarded-For") or "").strip()
    if forwarded:
        return forwarded.split(",")[0].strip()
    return (request.remote_addr or "").strip()


def _ensure_presence_session(user: Optional[User]) -> None:
    if not session.get("authenticated"):
        return
    actor = (session.get("actor") or "").strip()
    if not actor:
        return
    now = _utcnow()
    presence_id = session.get("presence_session_id")
    with db_session() as db:
        row = None
        if presence_id:
            try:
                row = db.get(UserPresenceSession, int(presence_id))
            except (TypeError, ValueError):
                row = None
            if row and row.logout_at is None:
                row.last_seen_at = now
                return
        token = secrets.token_hex(16)
        presence = UserPresenceSession(
            user_id=user.id if user else None,
            username=actor,
            session_token=token,
            ip_address=_client_ip(),
            user_agent=(request.headers.get("User-Agent") or "")[:512],
            login_at=now,
            last_seen_at=now,
            logout_at=None,
            ended_reason="",
        )
        db.add(presence)
        db.flush()
        session["presence_session_id"] = int(presence.id)


def _close_presence_session(reason: str = "logout") -> None:
    sid = session.get("presence_session_id")
    if not sid:
        return
    try:
        sid_int = int(sid)
    except (TypeError, ValueError):
        return
    with db_session() as db:
        row = db.get(UserPresenceSession, sid_int)
        if row and row.logout_at is None:
            row.logout_at = _utcnow()
            row.ended_reason = reason


DISCUSSION_THREAD_STATUSES = {"open", "needs_author_response", "resolved"}


def _serialize_discussion_thread(thread: ChangeDiscussionThread) -> Dict[str, Any]:
    return {
        "id": thread.id,
        "change_request_id": thread.change_request_id,
        "subject": thread.subject,
        "status": thread.status,
        "requires_resolution": bool(thread.requires_resolution),
        "created_by": thread.created_by,
        "created_role": thread.created_role,
        "created_at": thread.created_at.isoformat() if thread.created_at else None,
        "updated_at": thread.updated_at.isoformat() if thread.updated_at else None,
        "resolved_by": thread.resolved_by,
        "resolved_at": thread.resolved_at.isoformat() if thread.resolved_at else None,
        "messages": [
            {
                "id": m.id,
                "author": m.author,
                "author_role": m.author_role,
                "body": m.body,
                "kind": m.kind,
                "created_at": m.created_at.isoformat() if m.created_at else None,
            }
            for m in sorted(thread.messages or [], key=lambda x: x.id or 0)
        ],
    }


def _extract_mentions(text_value: str) -> List[str]:
    return sorted(set(re.findall(r"@([A-Za-z0-9_.-]+)", text_value or "")))


def _active_users_by_role(session, role: str) -> List[User]:
    return session.execute(select(User).where(User.role == role, User.is_active == True)).scalars().all()


def _active_user_by_username(session, username: str) -> Optional[User]:
    uname = (username or "").strip()
    if not uname:
        return None
    return session.execute(select(User).where(User.username == uname, User.is_active == True)).scalars().first()


def _render_notification(event_type: str, payload: Dict[str, Any]) -> Tuple[str, str]:
    p = payload or {}
    if event_type == "cr_submitted":
        return (
            f"[de_matrix] Change #{p.get('change_id')} submitted for review",
            (
                f"Change request #{p.get('change_id')}\n"
                f"Title: {p.get('title')}\n"
                f"Author: {p.get('author')}\n"
                f"Status: {p.get('status')}\n"
                "Please review in /changes."
            ),
        )
    if event_type == "cr_status":
        return (
            f"[de_matrix] Change #{p.get('change_id')} status -> {p.get('status')}",
            (
                f"Change request #{p.get('change_id')}\n"
                f"Title: {p.get('title')}\n"
                f"New status: {p.get('status')}\n"
                f"Changed by: {p.get('actor')}\n"
                f"Comment: {p.get('comment') or '-'}"
            ),
        )
    if event_type == "mention":
        return (
            f"[de_matrix] Mention in discussion for change #{p.get('change_id')}",
            (
                f"You were mentioned by {p.get('actor')} in discussion of change #{p.get('change_id')}.\n\n"
                f"{p.get('text') or ''}"
            ),
        )
    if event_type == "release":
        return (
            f"[de_matrix] New release deployed: {p.get('ref')}",
            f"A new release has been deployed.\n\nRef: {p.get('ref')}\nOpen the application to review latest changes.",
        )
    return (
        f"[de_matrix] Notification: {event_type}",
        str(p),
    )


def _create_notification_log(
    session,
    *,
    event_type: str,
    created_by: str,
    recipients: List[str],
    subject: str,
    body: str,
    context: Optional[Dict[str, Any]] = None,
) -> NotificationLog:
    log = NotificationLog(
        event_type=event_type,
        status="pending",
        subject=subject,
        body=body,
        recipients=recipients or [],
        context=context or {},
        error="",
        attempts=0,
        created_by=created_by or "system",
    )
    session.add(log)
    session.flush()
    return log


def _deliver_notification_log(log: NotificationLog) -> Tuple[bool, str]:
    if not NOTIFICATIONS_ENABLED:
        log.status = "skipped"
        log.error = "notifications disabled"
        log.attempts = (log.attempts or 0) + 1
        log.last_attempt_at = datetime.now(timezone.utc)
        return False, "notifications disabled"
    to_list = sorted({(r or "").strip() for r in (log.recipients or []) if (r or "").strip()})
    if not to_list:
        log.status = "skipped"
        log.error = "no recipients"
        log.attempts = (log.attempts or 0) + 1
        log.last_attempt_at = datetime.now(timezone.utc)
        return False, "no recipients"
    msg = EmailMessage()
    msg["From"] = SMTP_FROM
    msg["To"] = ", ".join(to_list)
    msg["Subject"] = log.subject
    msg.set_content(log.body)
    log.attempts = (log.attempts or 0) + 1
    log.last_attempt_at = datetime.now(timezone.utc)
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as smtp:
            smtp.send_message(msg)
        log.status = "sent"
        log.error = ""
        log.sent_at = datetime.now(timezone.utc)
        return True, ""
    except Exception as exc:
        err = str(exc)
        log.status = "failed"
        log.error = err
        print(f"[notify] send failed: {err}")
        return False, err


def _notify_cr_submitted(session, cr: ChangeRequest) -> None:
    admins = _active_users_by_role(session, "admin")
    recipients = [u.email for u in admins if u.email and u.username != cr.created_by]
    subject, body = _render_notification(
        "cr_submitted",
        {
            "change_id": cr.id,
            "title": cr.title,
            "author": cr.created_by,
            "status": cr.status,
        },
    )
    log = _create_notification_log(
        session,
        event_type="cr_submitted",
        created_by=cr.created_by,
        recipients=recipients,
        subject=subject,
        body=body,
        context={"change_id": cr.id},
    )
    _deliver_notification_log(log)


def _notify_cr_status_to_author(session, cr: ChangeRequest, status: str, actor: str, comment: str = "") -> None:
    author = _active_user_by_username(session, cr.created_by)
    recipients = [author.email] if author and author.email else []
    subject, body = _render_notification(
        "cr_status",
        {
            "change_id": cr.id,
            "title": cr.title,
            "status": status,
            "actor": actor,
            "comment": comment,
        },
    )
    log = _create_notification_log(
        session,
        event_type="cr_status",
        created_by=actor,
        recipients=recipients,
        subject=subject,
        body=body,
        context={"change_id": cr.id, "status": status},
    )
    _deliver_notification_log(log)


def _notify_mentions(session, change_id: int, actor: str, text_value: str) -> None:
    usernames = _extract_mentions(text_value or "")
    if not usernames:
        return
    recipients: List[str] = []
    for uname in usernames:
        u = _active_user_by_username(session, uname)
        if u and u.email and u.username != actor:
            recipients.append(u.email)
    subject, body = _render_notification(
        "mention",
        {
            "change_id": change_id,
            "actor": actor,
            "text": text_value,
        },
    )
    log = _create_notification_log(
        session,
        event_type="mention",
        created_by=actor,
        recipients=recipients,
        subject=subject,
        body=body,
        context={"change_id": change_id, "mentions": usernames},
    )
    _deliver_notification_log(log)


_SAFE_IDENT_RE = re.compile(r"^[A-Za-z_]\w*$")


def _quote_ident(name: str) -> str:
    if not _SAFE_IDENT_RE.match(name or ""):
        raise ValueError(f"Unsafe identifier: {name}")
    return f'"{name}"'


def _qualified_ident(schema: str, name: str) -> str:
    return f"{_quote_ident(schema)}.{_quote_ident(name)}"


def _build_table_ddl(session, schema: str, name: str) -> str:
    cols = session.execute(
        text(
            """
            SELECT
                c.column_name,
                c.data_type,
                c.is_nullable,
                c.column_default
            FROM information_schema.columns c
            WHERE c.table_schema = :schema AND c.table_name = :name
            ORDER BY c.ordinal_position
            """
        ),
        {"schema": schema, "name": name},
    ).mappings().all()
    if not cols:
        return f"-- No columns found for {schema}.{name}"
    constraints = session.execute(
        text(
            """
            SELECT
                tc.constraint_name,
                tc.constraint_type,
                pg_get_constraintdef(pg_constraint.oid) AS constraint_def
            FROM information_schema.table_constraints tc
            JOIN pg_namespace ns ON ns.nspname = tc.table_schema
            JOIN pg_class cls ON cls.relname = tc.table_name AND cls.relnamespace = ns.oid
            JOIN pg_constraint ON pg_constraint.conrelid = cls.oid
                AND pg_constraint.conname = tc.constraint_name
            WHERE tc.table_schema = :schema AND tc.table_name = :name
            ORDER BY tc.constraint_name
            """
        ),
        {"schema": schema, "name": name},
    ).mappings().all()
    lines = [f"CREATE TABLE {_qualified_ident(schema, name)} ("]
    column_lines = []
    for col in cols:
        line = f"    {_quote_ident(col['column_name'])} {col['data_type']}"
        if col["column_default"] is not None:
            line += f" DEFAULT {col['column_default']}"
        if col["is_nullable"] == "NO":
            line += " NOT NULL"
        column_lines.append(line)
    for c in constraints:
        c_line = f"    CONSTRAINT {_quote_ident(c['constraint_name'])} {c['constraint_def']}"
        column_lines.append(c_line)
    lines.append(",\n".join(column_lines))
    lines.append(");")
    return "\n".join(lines)


def _format_db_error(exc: Exception) -> Dict:
    out = {
        "message": str(exc),
        "type": exc.__class__.__name__,
    }
    orig = getattr(exc, "orig", None)
    if orig is not None:
        diag = getattr(orig, "diag", None)
        if diag is not None:
            primary = getattr(diag, "message_primary", None)
            detail = getattr(diag, "message_detail", None)
            position = getattr(diag, "statement_position", None)
            if primary:
                out["primary"] = primary
            if detail:
                out["detail"] = detail
            if position:
                out["position"] = position
    return out


def _collect_template_ids_from_nodes(nodes: List[Dict[str, Any]]) -> Set[str]:
    out: Set[str] = set()
    for n in nodes or []:
        if not isinstance(n, dict):
            continue
        tid = str(n.get("template_id") or "").strip()
        if tid:
            out.add(tid)
        out |= _collect_template_ids_from_nodes(n.get("children") or [])
    return out


def _build_tree_edit_warnings(current_unified: Dict[str, Any], edited_nodes: List[Dict[str, Any]]) -> Dict[str, Any]:
    current_templates = _collect_template_ids_from_nodes(current_unified.get("nodes") or [])
    edited_templates = _collect_template_ids_from_nodes(edited_nodes or [])
    removed_templates = sorted(current_templates - edited_templates)
    template_defs = current_unified.get("action_templates") or {}
    removed_literature: Set[str] = set()
    for tid in removed_templates:
        t = template_defs.get(tid) or {}
        for lit_id in t.get("resource_ids") or []:
            if lit_id:
                removed_literature.add(str(lit_id))
    return {
        "removed_template_ids": removed_templates,
        "removed_template_count": len(removed_templates),
        "affected_literature_ids": sorted(removed_literature),
        "affected_literature_count": len(removed_literature),
    }


def _literature_dir():
    """Каталог для скачанной литературы (data/library)."""
    cfg = get_meta()
    rel = cfg.get('literature_dir') or cfg.get('library_dir') or 'data/library'
    path = os.path.join(BASE_DIR, rel) if not os.path.isabs(rel) else rel
    os.makedirs(path, exist_ok=True)
    return path


def _invalidate_caches():
    """Сброс кэшей: при перезапуске или изменении источника данные перезагружаются (autoscale)."""
    global _matrix, _meta, _tree, _current_source_file
    _matrix = None
    _meta = None
    _tree = None
    _current_source_file = None
    invalidate_metadata_cache()


def _ensure_db_schema():
    global _admin_seed_checked
    if _is_db_mode():
        if ENGINE.dialect.name == "postgresql":
            with ENGINE.begin() as conn:
                conn.execute(text(f'CREATE SCHEMA IF NOT EXISTS "{MATRIX_STRUCT_SCHEMA}"'))
        Base.metadata.create_all(bind=ENGINE)
        with db_session() as session:
            session.execute(
                text(
                    "ALTER TABLE users "
                    "ADD COLUMN IF NOT EXISTS updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()"
                    )
                )
        if not _admin_seed_checked:
            _ensure_admin_seed()
            _admin_seed_checked = True


def _path_config():
    """Только настройки путей (без мета из источника). Используется до загрузки данных."""
    return load_app_config()


def _source_dir_path():
    """Абсолютный путь к каталогу источников (source_dir из настроек)."""
    cfg = _meta if _meta is not None else _path_config()
    rel = cfg.get("source_dir") or "data/sources"
    path = os.path.join(BASE_DIR, rel) if not os.path.isabs(rel) else rel
    return path


def _checkpoint_path():
    """Абсолютный путь к файлу чекпоинта."""
    cfg = _meta if _meta is not None else _path_config()
    rel = cfg.get("checkpoint_file") or "data/checkpoint.yaml"
    path = os.path.join(BASE_DIR, rel) if not os.path.isabs(rel) else rel
    return path


def _list_source_files():
    """Список имён файлов-источников в source_dir (JSON, YAML, Excel)."""
    allowed = (".json", ".yaml", ".yml", ".xlsx", ".xls")
    src_dir = _source_dir_path()
    if not os.path.isdir(src_dir):
        return []
    return sorted(
        f for f in os.listdir(src_dir)
        if os.path.isfile(os.path.join(src_dir, f)) and f.lower().endswith(allowed)
    )


def _current_source_for_backup() -> str:
    """Имя текущего файла-источника для бэкапа (чекпоинт или default)."""
    global _current_source_file
    if _is_db_mode():
        return ""
    if _current_source_file:
        return _current_source_file
    cfg = _path_config()
    checkpoint = load_checkpoint(PathLib(_checkpoint_path()))
    if checkpoint and checkpoint.get("source_file"):
        return checkpoint["source_file"]
    default = cfg.get("default_source")
    candidates = _list_source_files()
    if default and default in candidates:
        return default
    return candidates[0] if candidates else ""


def _create_version_backup(change_type: str, note: str = "") -> str:
    """
    Создаёт версионный бэкап перед изменением.
    Также гарантирует наличие stable-состояния.
    """
    if _is_db_mode():
        return ""
    source_file = _current_source_for_backup()
    if not source_file:
        return ""
    base = PathLib(BASE_DIR)
    cfg = base / "config"
    src = PathLib(_source_dir_path())
    cp = _checkpoint_path()
    ensure_stable_backup(
        base_dir=base,
        config_dir=cfg,
        source_dir=src,
        source_filename=source_file,
        checkpoint_path=cp,
    )
    backup_id = create_backup(
        base_dir=base,
        config_dir=cfg,
        source_dir=src,
        source_filename=source_file,
        checkpoint_path=cp,
        change_type=change_type,
        note=note,
    )
    return backup_id or ""


def _ensure_data_loaded(force_source_filename: str = None):
    """
    Загружает матрицу и мета-ключи из PostgreSQL (единый unified-снимок).
    Параметр force_source_filename зарезервирован для совместимости вызовов и игнорируется в DB-режиме.
    """
    global _matrix, _tree, _meta, _current_source_file
    if _matrix is not None and _tree is not None and not force_source_filename:
        return
    if not _is_db_mode():
        raise RuntimeError("Runtime is DB-only; file source loading is disabled")
    _ensure_db_schema()
    path_cfg = _path_config()
    with db_session() as session:
        unified = load_unified_from_db(session, literature=load_literature_map())
    nodes = unified.get("nodes") or []
    _matrix = {"nodes": nodes, "domains": []}
    _tree = build_tree_from_matrix_data(_matrix)
    meta_from_source = {k: unified.get(k, {} if k != "action_examples" else []) for k in META_KEYS}
    _meta = {**path_cfg, **meta_from_source}
    _current_source_file = "db://postgres"

# ----- Вспомогательные функции для генерации иконок и цветов -----
DOMAIN_ICONS = [
    "database", "cloud", "code-branch", "check-circle", "users",
    "server", "chart-line", "cube", "cogs", "file-alt"
]
SKILL_ICONS = [
    "cube", "bolt", "layer-group", "fire", "clock", "python",
    "sync-alt", "chart-line", "check-double", "project-diagram", "file-alt",
    "comments", "tasks", "chalkboard-teacher"
]

def get_domain_icon(domain_name, index):
    return DOMAIN_ICONS[index % len(DOMAIN_ICONS)]

def get_skill_icon(skill_name, index):
    return SKILL_ICONS[index % len(SKILL_ICONS)]

def string_to_hsl(text, s=70, l=60):
    hash_obj = hashlib.md5(text.encode())
    hue = int(hash_obj.hexdigest()[:6], 16) % 360
    return f"hsl({hue}, {s}%, {l}%)"

def get_domain_color(domain_name):
    return string_to_hsl(domain_name, s=70, l=60)

def get_skill_color(skill_name, domain_color, skill_index):
    match = re.match(r'hsl\((\d+), (\d+)%, (\d+)%\)', domain_color)
    if match:
        hue = (int(match.group(1)) + skill_index * 30) % 360
        s = int(match.group(2))
        l = int(match.group(3))
        return f"hsl({hue}, {s}%, {l}%)"
    return domain_color

# ----------------------------------------------------------------

def load_json(path, default):
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Файл {path} не найден. Используется значение по умолчанию.")
        return default
    except Exception as e:
        print(f"Ошибка загрузки {path}: {e}")
        return default

def get_meta():
    """Конфиг: пути из settings + данные из источника (текст) + стили/инструменты из config/metadata.yaml."""
    global _meta
    if _meta is None:
        _meta = _path_config()
        _ensure_data_loaded()
    metadata = load_metadata()
    out = {
        **_meta,
        "stack_labels": metadata.get("stack_labels", {}),
        "tools_patterns": metadata.get("tools_patterns", {}),
        "tools_groups": metadata.get("tools_groups", {}),
    }
    out.setdefault("action_templates", {})
    out.setdefault("literature", {})
    out.setdefault("action_examples", [])
    out.setdefault("ui_config", {})
    return out


def get_matrix():
    """Структура матрицы из чекпоинта или файла-источника (source_dir). Сверка по хешу при каждой загрузке."""
    _ensure_data_loaded()
    return _matrix if _matrix is not None else {"nodes": [], "domains": []}


def _matrix_root_count(matrix: Optional[Dict]) -> int:
    m = matrix or {}
    return len(m.get("nodes") or [])


def _parse_export_domain_idxs(raw: str) -> List[int]:
    idxs: List[int] = []
    for part in (raw or "").strip().split(","):
        part = part.strip()
        if part.isdigit():
            idxs.append(int(part))
    return idxs


def get_tree():
    """Дерево матрицы (autoscale по структуре источника). Листья — узлы без children."""
    _ensure_data_loaded()
    return _tree if _tree is not None else []


def _count_leaves_under_node(node: Dict) -> int:
    """Число листьев в поддереве generic-узла (для сайдбара)."""
    if not isinstance(node, dict):
        return 0
    ch = node.get("children") or []
    if not ch:
        return 1
    n = 0
    for c in ch:
        if isinstance(c, dict):
            n += _count_leaves_under_node(c)
    return n


def _build_domain_graph_from_generic_tree(domain_idx: int, meta: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """
    Граф домена для режима matrix_nodes: обход get_tree()[domain_idx] (path уже проставлены).
    Листья — leaf_path для /leaf/...; типы узлов без легаси-имён: group / branch / leaf (не skill/action/subaction).
    """
    tree = get_tree()
    if domain_idx < 0 or domain_idx >= len(tree):
        return None
    root_tn = tree[domain_idx]
    if not isinstance(root_tn, dict):
        return None
    domain_name = root_tn.get("name", "") or ""
    domain_color = get_domain_color(domain_name)
    domain_icon = get_domain_icon(domain_name, domain_idx)
    domain_root_id = f"dg_root_{domain_idx}"
    nodes_out: List[Dict[str, Any]] = []
    links_out: List[Dict[str, str]] = []
    nodes_out.append(
        {
            "id": domain_root_id,
            "name": domain_name,
            "type": "domain_root",
            "color": domain_color,
            "icon": domain_icon,
            "level": 0,
            "description": str(root_tn.get("description") or "").strip(),
            "responsible": root_tn.get("responsible") or "",
            "level_tags": normalize_level_tags(root_tn.get("level_tags") or root_tn.get("level_tag")),
            "open_action_url": False,
        }
    )
    templates = meta.get("action_templates") or {}

    def visit(parent_gid: str, tnode: Dict[str, Any], parent_type: str) -> None:
        children = tnode.get("children") or []
        if not isinstance(children, list):
            return
        for child in children:
            if not isinstance(child, dict):
                continue
            cpath = child.get("path")
            if not isinstance(cpath, list) or not cpath:
                continue
            cid = "dg_n_" + "_".join(str(p) for p in cpath)
            chch = child.get("children") or []
            has_ch = isinstance(chch, list) and bool(chch)
            ln = len(cpath)
            if ln == 2:
                typ = "group"
                vis_level = 1
            elif not has_ch:
                typ = "leaf"
                vis_level = min(ln - 1, 3)
            else:
                typ = "branch"
                vis_level = min(ln - 1, 3)
            leaf_path_str: Optional[str] = None
            if not has_ch:
                leaf_path_str = "/".join(str(p) for p in cpath)
            name = child.get("name") or ""
            group_color_idx = int(cpath[1]) if ln > 1 else 0
            tpl_id = child.get("template_id")
            template = templates.get(tpl_id, {}) if tpl_id else {}
            enriched = enrich_action(child, template if isinstance(template, dict) else {}, meta)
            stack_labels = enriched.get("stack_labels") or []

            if parent_type == "domain_root":
                elbl = "содержит"
            elif parent_type == "group":
                elbl = "выполняет"
            else:
                elbl = "содержит"
            links_out.append({"source": parent_gid, "target": cid, "label": elbl})

            node_payload: Dict[str, Any] = {
                "id": cid,
                "name": name,
                "full_name": name,
                "type": typ,
                "level": vis_level,
                "path": list(cpath),
                "leaf_path": leaf_path_str,
                "description": str(child.get("description") or "").strip(),
                "responsible": child.get("responsible") or "",
                "level_tag": child.get("level_tag"),
                "level_tags": action_level_tags_for_json(child),
                "level_sticker": child.get("level_sticker") or "",
                "open_action_url": False,
            }
            if typ == "group":
                node_payload["color"] = get_skill_color(name, domain_color, group_color_idx)
                node_payload["icon"] = get_skill_icon(name, group_color_idx)
            else:
                node_payload["color"] = "#f39c12"
            nodes_out.append(node_payload)

            if stack_labels:
                for stack_idx, stack in enumerate(stack_labels):
                    stack_id = f"dg_stack_{'_'.join(str(p) for p in cpath)}_{stack_idx}"
                    if not any(n["id"] == stack_id for n in nodes_out):
                        nodes_out.append(
                            {
                                "id": stack_id,
                                "name": stack.get("name", stack.get("key", "Technology")),
                                "type": "stack",
                                "icon": stack.get("icon", "cube"),
                                "color": stack.get("color", "#9b59b6"),
                                "level": 4,
                                "description": stack.get("description", ""),
                                "open_action_url": False,
                            }
                        )
                    links_out.append({"source": cid, "target": stack_id, "label": "использует"})

            visit(cid, child, typ)

    visit(domain_root_id, root_tn, "domain_root")
    return {
        "domain": {"name": domain_name, "color": domain_color, "icon": domain_icon},
        "nodes": nodes_out,
        "links": links_out,
    }


def _build_global_graph_from_generic_tree(meta: Dict[str, Any]) -> Dict[str, Any]:
    """
    Общий граф (/graph) для режима matrix_nodes: все корни get_tree() под синтетическим root.
    Листья — leaf_path; типы group / branch / leaf (без skill/action/subaction в JSON).
    """
    tree = get_tree()
    nodes_out: List[Dict[str, Any]] = [
        {
            "id": "root",
            "name": "Middle Data Engineer",
            "type": "root",
            "level": 0,
            "open_action_url": False,
        }
    ]
    links_out: List[Dict[str, str]] = []
    templates = meta.get("action_templates") or {}

    for di, root_tn in enumerate(tree):
        if not isinstance(root_tn, dict):
            continue
        did = f"d{di}"
        domain_name = root_tn.get("name", "") or ""
        domain_color = get_domain_color(domain_name)
        domain_icon = get_domain_icon(domain_name, di)
        nodes_out.append(
            {
                "id": did,
                "name": domain_name,
                "type": "domain",
                "level": 1,
                "color": domain_color,
                "icon": domain_icon,
                "description": str(root_tn.get("description") or "").strip(),
                "responsible": root_tn.get("responsible") or "",
                "level_tags": normalize_level_tags(root_tn.get("level_tags") or root_tn.get("level_tag")),
                "open_action_url": False,
            }
        )
        links_out.append({"source": "root", "target": did})

        def visit(parent_id: str, tnode: Dict[str, Any], parent_type: str) -> None:
            children = tnode.get("children") or []
            if not isinstance(children, list):
                return
            for child in children:
                if not isinstance(child, dict):
                    continue
                cpath = child.get("path")
                if not isinstance(cpath, list) or not cpath:
                    continue
                cid = "g_" + "_".join(str(p) for p in cpath)
                chch = child.get("children") or []
                has_ch = isinstance(chch, list) and bool(chch)
                ln = len(cpath)
                if ln == 2:
                    typ = "group"
                elif not has_ch:
                    typ = "leaf"
                else:
                    typ = "branch"
                vis_level = ln
                leaf_path_str: Optional[str] = None
                if not has_ch:
                    leaf_path_str = "/".join(str(p) for p in cpath)
                name = child.get("name") or ""
                group_color_idx = int(cpath[1]) if ln > 1 else 0

                if parent_type == "domain":
                    elbl = "содержит"
                elif parent_type == "group":
                    elbl = "выполняет"
                else:
                    elbl = "содержит"
                links_out.append({"source": parent_id, "target": cid, "label": elbl})

                tpl_id = child.get("template_id")
                template = templates.get(tpl_id, {}) if tpl_id else {}
                enriched = enrich_action(child, template if isinstance(template, dict) else {}, meta)
                stack_labels = enriched.get("stack_labels") or []

                node_payload: Dict[str, Any] = {
                    "id": cid,
                    "name": name,
                    "full_name": name,
                    "type": typ,
                    "level": vis_level,
                    "path": list(cpath),
                    "leaf_path": leaf_path_str,
                    "description": str(child.get("description") or "").strip(),
                    "responsible": child.get("responsible") or "",
                    "level_tag": child.get("level_tag"),
                    "level_tags": action_level_tags_for_json(child),
                    "level_sticker": child.get("level_sticker") or "",
                    "open_action_url": False,
                }
                if typ == "group":
                    node_payload["color"] = get_skill_color(name, domain_color, group_color_idx)
                    node_payload["icon"] = get_skill_icon(name, group_color_idx)
                else:
                    node_payload["color"] = "#f39c12"
                nodes_out.append(node_payload)

                if stack_labels:
                    for stack_idx, stack in enumerate(stack_labels):
                        stack_id = f"gg_stack_{'_'.join(str(p) for p in cpath)}_{stack_idx}"
                        if not any(n["id"] == stack_id for n in nodes_out):
                            nodes_out.append(
                                {
                                    "id": stack_id,
                                    "name": stack.get("name", stack.get("key", "Technology")),
                                    "type": "stack",
                                    "icon": stack.get("icon", "cube"),
                                    "color": stack.get("color", "#9b59b6"),
                                    "level": ln + 2,
                                    "description": stack.get("description", ""),
                                    "open_action_url": False,
                                }
                            )
                        links_out.append({"source": cid, "target": stack_id, "label": "использует"})

                visit(cid, child, typ)

        visit(did, root_tn, "domain")

    return {"nodes": nodes_out, "links": links_out}


def _expected_leaf_paths_from_matrix(matrix: Dict) -> list:
    """Ожидаемые leaf-path по дереву nodes (для проверки автоскейла)."""
    tree = build_tree_from_matrix_data(matrix or {})
    leaves = collect_leaves(tree)
    return ["/".join(str(x) for x in (n.get("path") or [])) for n in leaves if n.get("path")]

def save_meta(meta_dict):
    """Сохраняет метаданные в единый источник (текущий файл в source_dir). Обновляет чекпоинт при следующей загрузке."""
    global _meta
    path_cfg = _path_config()
    _meta = {**path_cfg, **meta_dict}
    if not _is_db_mode():
        raise RuntimeError("Runtime is DB-only; file source saving is disabled")
    _ensure_db_schema()
    with db_session() as session:
        unified = load_unified_from_db(session, literature=meta_dict.get("literature", {}) or {})
        mx = get_matrix() or {}
        unified["domains"] = []
        unified["nodes"] = mx.get("nodes") or []
        for k in META_KEYS:
            unified[k] = meta_dict.get(k, {} if k != "action_examples" else [])
        replace_unified_in_db(session, unified)
    for lit_id, item in (meta_dict.get("literature") or {}).items():
        upsert_literature_item(str(lit_id), item or {})
    _invalidate_caches()

def slugify(text):
    if not text:
        return ''
    return re.sub(r'[^\w\s-]', '', text.lower()).strip().replace(' ', '-')

def find_free_port(start_port=5000, max_attempts=10):
    for port in range(start_port, start_port + max_attempts):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            try:
                s.bind(('0.0.0.0', port))
                return port
            except OSError:
                continue
    return None

def enrich_action(action_obj, template, meta):
    """Обогащает действие: стек по stack_refs из шаблона (стили из config), инструменты — по паттернам в тексте."""
    stack_labels = meta.get("stack_labels", {})
    action_examples = meta.get("action_examples", [])
    literature = meta.get("literature", {})
    tools_patterns = meta.get("tools_patterns", {})
    tools_groups = meta.get("tools_groups", {})

    if not template:
        return {
            "stack_labels": [],
            "tools": [],
            "examples": [],
            "literature": [],
        }

    stack_refs = template.get("stack_refs", [])
    resolved_stack = [{"key": ref, **stack_labels[ref]} for ref in stack_refs if ref in stack_labels]

    leaf_text = (action_obj.get("text") or action_obj.get("name") or "")
    template_text = " ".join(template.get("minimal_requirements") or []) + " " + " ".join(template.get("antipatterns") or [])
    resolved_tools = get_tools_for_text(leaf_text, template_text, tools_patterns, tools_groups)

    examples_refs = template.get("examples_refs", [])
    examples_by_id = {e['id']: e for e in action_examples if 'id' in e}
    resolved_examples = []
    for ref in examples_refs:
        if ref in examples_by_id:
            ex = examples_by_id[ref]
            code = ex.get('code', '').replace('<', '&lt;').replace('>', '&gt;')
            resolved_examples.append({
                'html': f'<pre><code class="language-{ex.get("language", "sql")}">{code}</code></pre>',
                'title': ex.get('title', '')
            })
    
    resource_ids = template.get('resource_ids', [])
    resolved_literature = []
    for rid in resource_ids:
        if rid in literature:
            resolved_literature.append({"id": rid, **literature[rid]})
    
    return {
        'stack_labels': resolved_stack,
        'tools': resolved_tools,
        'examples': resolved_examples,
        'literature': resolved_literature
    }

def build_description(action_obj, template, domain, skill, meta):
    minimal = template.get('minimal_requirements', [])
    antipatterns = template.get('antipatterns', [])
    ui = meta.get('ui_config', {})
    titles = ui.get('section_titles', {})

    node_note = (action_obj.get("description") or "").strip()
    note_html = ""
    if node_note:
        note_esc = html_lib.escape(node_note).replace("\n", "<br>\n")
        # Текст узла — полное описание; заголовок «Описание» задаётся в шаблоне модалки/страницы.
        note_html = f'<div class="matrix-node-own-description">{note_esc}</div>'

    if not minimal and not antipatterns:
        body = f"""
            <h4>📋 Действие в контексте {html_lib.escape(domain.get('name') or '')}</h4>
            <p><strong>{html_lib.escape(action_obj.get('text') or '')}</strong> относится к навыку <strong>{html_lib.escape(skill.get('name') or '')}</strong>.</p>
        """
        if node_note:
            return note_html + body
        return body + "<p>Описание пока не добавлено.</p>"

    html = note_html + f"<h4>📋 {html_lib.escape(template.get('name', 'Действие'))}</h4>"
    if minimal:
        html += f"<h5>✅ {titles.get('minimal_requirements', 'Минимальный объем')}:</h5><ul>"
        for req in minimal:
            html += f"<li>{req}</li>"
        html += "</ul>"
    if antipatterns:
        html += f"<h5>⚠️ {titles.get('antipatterns', 'Антипаттерны')}:</h5><ul>"
        for anti in antipatterns:
            html += f"<li>{anti}</li>"
        html += "</ul>"
    html += f"<p><small>Контекст: <strong>{html_lib.escape(domain.get('name') or '')}</strong> → <strong>{html_lib.escape(skill.get('name') or '')}</strong></small></p>"
    return html

def resolve_leaf_by_path(path_str):
    """
    По path (например "0/1/2" или "0/1/2/0") возвращает (domain, skill, action, parent_action_text)
    для рендера страницы листа. action — dict с text, template_id; domain/skill — dict с name, description у skill.
    Если узел не найден или не лист — возвращает None.
    """
    try:
        path = [int(x) for x in path_str.strip("/").split("/") if x.strip()]
    except (ValueError, AttributeError):
        return None
    if not path:
        return None
    tree = get_tree()
    node = get_node_by_path(tree, path)
    if not node or node.get("children"):
        return None
    ancestors = get_ancestors(tree, path)
    empty_skill = {
        "name": "",
        "description": "",
        "responsible": "",
        "level_sticker": "",
    }
    if len(ancestors) >= 2:
        domain = {"name": ancestors[0].get("name", "")}
        skill = {
            "name": ancestors[1].get("name", ""),
            "description": ancestors[1].get("description", ""),
            "responsible": ancestors[1].get("responsible", ""),
            "level_sticker": ancestors[1].get("level_sticker", ""),
        }
    elif len(ancestors) == 1:
        domain = {"name": ancestors[0].get("name", "")}
        skill = dict(empty_skill)
    else:
        domain = {"name": ""}
        skill = dict(empty_skill)
    action = {
        "text": node.get("name", ""),
        "template_id": node.get("template_id"),
        "level_tag": node.get("level_tag"),
        "level_tags": action_level_tags_for_json(node),
        "leaf_view": dict(node.get("leaf_view") or {}),
        "review_questions": node.get("review_questions", []),
    }
    nd = (node.get("description") or "").strip()
    if nd:
        action["description"] = nd
    # Подуровень: у листа есть минимум два предка над «действием» (домен → навык → … → родитель листа).
    parent_action_text = ancestors[-1].get("name", "") if len(ancestors) >= 3 else None
    return (domain, skill, action, parent_action_text)

# РЕГИСТРАЦИЯ ФИЛЬТРОВ JINJA2
app.jinja_env.filters['slugify'] = slugify

@app.template_filter('domain_icon')
def domain_icon_filter(domain_name, index):
    return get_domain_icon(domain_name, index)

@app.template_filter('skill_icon')
def skill_icon_filter(skill_name, index):
    return get_skill_icon(skill_name, index)

@app.template_filter('domain_color')
def domain_color_filter(domain_name):
    return get_domain_color(domain_name)

@app.template_filter('skill_color')
def skill_color_filter(skill_name, domain_color, index):
    return get_skill_color(skill_name, domain_color, index)


@app.template_filter("matrix_schema_ui")
def matrix_schema_ui_filter(schema):
    """Схема колонок для UI: без суффиксов (item)/(leaf_view) в header/label."""
    return schema_entries_for_ui(schema if isinstance(schema, list) else [])


@app.context_processor
def inject_globals():
    current_actor, current_role = _extract_actor_role()
    sidebar_badges = {
        "changes_pending": 0,
        "notifications_unsent": 0,
        "presence_online": 0,
        "online_usernames": [],
    }
    if bool(session.get("authenticated")):
        try:
            _ensure_db_schema()
            pending_statuses = ("draft", "submitted", "in_review", "approved")
            online_cutoff = _utcnow() - timedelta(seconds=PRESENCE_ONLINE_SECONDS)
            with db_session() as db:
                sidebar_badges["changes_pending"] = int(
                    db.execute(
                        select(func.count(ChangeRequest.id)).where(ChangeRequest.status.in_(pending_statuses))
                    ).scalar()
                    or 0
                )
                online_rows = db.execute(
                    select(func.distinct(UserPresenceSession.username))
                    .select_from(UserPresenceSession)
                    .join(User, User.username == UserPresenceSession.username)
                    .where(
                        User.is_active == True,
                        UserPresenceSession.logout_at.is_(None),
                        UserPresenceSession.last_seen_at.is_not(None),
                        UserPresenceSession.last_seen_at >= online_cutoff,
                        UserPresenceSession.username.is_not(None),
                        UserPresenceSession.username != "",
                    )
                    .order_by(UserPresenceSession.username.asc())
                ).scalars().all()
                sidebar_badges["online_usernames"] = [u for u in online_rows if u]
                sidebar_badges["presence_online"] = len(sidebar_badges["online_usernames"])
                if current_role == "admin":
                    sidebar_badges["notifications_unsent"] = int(
                        db.execute(
                            select(func.count(NotificationLog.id)).where(NotificationLog.status.in_(["failed", "skipped"]))
                        ).scalar()
                        or 0
                    )
        except Exception:
            sidebar_badges = {
                "changes_pending": 0,
                "notifications_unsent": 0,
                "presence_online": 0,
                "online_usernames": [],
            }
    matrix = get_matrix() or {}
    nodes = matrix.get("nodes") or []
    sidebar_domains = []
    for i, root in enumerate(nodes):
        if not isinstance(root, dict):
            continue
        domain_color = get_domain_color(root.get("name", ""))
        skills_list = []
        for si, s in enumerate(root.get("children") or []):
            if not isinstance(s, dict):
                continue
            skills_list.append(
                {
                "name": s.get("name", ""),
                "index": si,
                "color": get_skill_color(s.get("name", ""), domain_color, si),
                "icon": get_skill_icon(s.get("name", ""), si),
                    "actions_count": _count_leaves_under_node(s),
                }
            )
        sidebar_domains.append(
            {
                "name": root.get("name", ""),
            "index": i,
            "color": domain_color,
                "icon": get_domain_icon(root.get("name", ""), i),
            "skills_count": len(skills_list),
            "skills": skills_list,
            }
        )
    _meta_ctx = get_meta()
    _ui_ctx = _meta_ctx.get("ui_config") or {}
    return {
        'ui_config': _ui_ctx,
        'unified_column_schema': schema_entries_for_ui(
            effective_matrix_column_schema(_ui_ctx), _ui_ctx
        ),
        'sidebar_domains': sidebar_domains,
        'get_domain_icon': get_domain_icon,
        'get_skill_icon': get_skill_icon,
        'get_domain_color': get_domain_color,
        'get_skill_color': get_skill_color,
        'author_name': AUTHOR_NAME,
        'author_telegram': AUTHOR_TELEGRAM,
        'repository_url': REPOSITORY_URL,
        'current_actor': current_actor,
        'current_role': current_role,
        'is_authenticated': bool(session.get("authenticated")),
        'sidebar_badges': sidebar_badges,
        'system_admin_username': ADMIN_USERNAME,
        'e2e_admin_username': E2E_ADMIN_USERNAME,
    }

# ----- ОСНОВНЫЕ МАРШРУТЫ -----


@app.route('/login', methods=['GET', 'POST'])
def login():
    _ensure_db_schema()
    next_url = (request.args.get("next") or request.form.get("next") or "/").strip() or "/"
    if session.get("authenticated") and _is_safe_next_url(next_url):
        return redirect(next_url)
    if request.method == 'GET':
        return render_template('login.html', next_url=next_url, auth_error=None)

    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""
    if not username or not password:
        return render_template('login.html', next_url=next_url, auth_error="Введите логин и пароль"), 400

    user = _load_user(username)
    if not user or not user.is_active or not user.password_hash:
        return render_template('login.html', next_url=next_url, auth_error="Неверный логин или пароль"), 401
    if not check_password_hash(user.password_hash, password):
        return render_template('login.html', next_url=next_url, auth_error="Неверный логин или пароль"), 401

    session["authenticated"] = True
    session["actor"] = username
    session["role"] = user.role
    session["must_change_password"] = bool(user.must_change_password)
    _ensure_presence_session(user)
    if user.must_change_password and request.path != "/account/password":
        return redirect(url_for("account_password"))
    if not _is_safe_next_url(next_url):
        next_url = "/"
    return redirect(next_url)


@app.route('/logout', methods=['GET'])
def logout():
    _close_presence_session("logout")
    session.clear()
    return redirect(url_for('login'))


@app.before_request
def enforce_authentication():
    if not AUTH_REQUIRED:
        return None
    if request.path.startswith("/static/") or request.path.startswith("/library/"):
        return None
    if request.path in ("/api/schema",):
        return None
    if request.path in ("/login", "/logout"):
        return None
    if session.get("authenticated"):
        return None
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": "Authentication required"}), 401
    return redirect(url_for("login", next=request.path))


@app.before_request
def enforce_password_change():
    if not session.get("authenticated"):
        return None
    must_change = bool(session.get("must_change_password"))
    if not must_change:
        return None
    allowed = {
        "account_password",
        "logout",
        "login",
        "static",
        "proxy_health",
    }
    endpoint = request.endpoint or ""
    if endpoint in allowed or endpoint.startswith("api_schema"):
        return None
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": "Password change required"}), 403
    return redirect(url_for("account_password"))


@app.before_request
def track_presence():
    if not session.get("authenticated"):
        return None
    if request.path.startswith("/static/") or request.path.startswith("/library/"):
        return None
    user = _load_user((session.get("actor") or "").strip())
    _ensure_presence_session(user)
    return None


@app.route('/account', methods=['GET'])
def account_page():
    auth, auth_err = _require_authenticated()
    if auth_err:
        return redirect(url_for("login", next=request.path))
    user = _load_user(auth["actor"])
    if not user:
        session.clear()
        return redirect(url_for("login", next=request.path))
    return render_template("account.html", profile={
        "username": user.username,
        "full_name": user.full_name or "",
        "email": user.email or "",
        "role": user.role,
        "must_change_password": bool(user.must_change_password),
    })


@app.route('/account/profile', methods=['POST'])
def account_update_profile():
    auth, auth_err = _require_authenticated()
    if auth_err:
        return auth_err
    data = request.get_json(silent=True) or request.form
    full_name = (data.get("full_name") or "").strip()
    email = (data.get("email") or "").strip()
    if not _is_valid_email_mask(email):
        return jsonify({"ok": False, "error": "Invalid email"}), 400
    with db_session() as db:
        user = db.execute(select(User).where(User.username == auth["actor"])).scalars().first()
        if not user:
            return jsonify({"ok": False, "error": "User not found"}), 404
        if _is_e2e_admin_username(user.username):
            return jsonify({"ok": False, "error": "e2e_admin profile cannot be edited"}), 403
        changed = False
        if not _is_system_admin_username(user.username):
            if not full_name:
                return jsonify({"ok": False, "error": "Full name cannot be empty"}), 400
            if user.full_name != full_name:
                user.full_name = full_name
                changed = True
        if user.email != email:
            user.email = email
            changed = True
        if changed:
            user.updated_at = _utcnow()
    return jsonify({"ok": True})


@app.route('/account/password', methods=['GET', 'POST'])
def account_password():
    auth, auth_err = _require_authenticated()
    if auth_err:
        return redirect(url_for("login", next=request.path))
    if request.method == "GET":
        return render_template("account_password.html", auth_error=None, must_change=bool(session.get("must_change_password")))

    current_password = request.form.get("current_password") or ""
    new_password = request.form.get("new_password") or ""
    confirm_password = request.form.get("confirm_password") or ""
    if len(new_password) < 10:
        return render_template("account_password.html", auth_error="Новый пароль слишком короткий (минимум 10 символов)", must_change=True), 400
    if new_password != confirm_password:
        return render_template("account_password.html", auth_error="Подтверждение пароля не совпадает", must_change=True), 400

    with db_session() as db:
        user = db.execute(select(User).where(User.username == auth["actor"])).scalars().first()
        if not user:
            session.clear()
            return redirect(url_for("login", next="/"))
        if not check_password_hash(user.password_hash, current_password):
            return render_template("account_password.html", auth_error="Текущий пароль неверный", must_change=bool(user.must_change_password)), 401
        user.password_hash = generate_password_hash(new_password)
        user.must_change_password = False

    session["must_change_password"] = False
    return redirect(url_for("account_page"))


@app.route('/')
def index():
    """Стартовая страница — дашборд."""
    matrix = get_matrix() or {}
    nodes = matrix.get("nodes") or []
    tree = build_tree_from_matrix_data({"nodes": nodes})
    total_skills = sum(len(n.get("children") or []) for n in nodes if isinstance(n, dict))
    total_leaves = len(collect_leaves(tree))
    domains_for_tpl = [{"name": (n.get("name") or "").strip()} for n in nodes[:5] if isinstance(n, dict)]
    return render_template(
        "home.html",
        domains=domains_for_tpl,
        stats={
            "domains": len(nodes),
        "skills": total_skills,
            "actions": total_leaves,
        },
    )


@app.route('/matrix')
def matrix_view():
    """Матрица — сетка карточек доменов."""
    _ensure_data_loaded()
    m = get_matrix() or {}
    has_data = bool(m.get("nodes"))
    return render_template(
        "matrix.html",
        domains=[],
        matrix_has_data=has_data,
    )


@app.route('/domain/<int:domain_idx>')
def domain_view(domain_idx):
    """Вью домена: дерево элементов слева направо."""
    matrix = get_matrix()
    nodes = (matrix or {}).get("nodes") or []
    if domain_idx < 0 or domain_idx >= _matrix_root_count(matrix):
        return render_template('404.html'), 404
    root = nodes[domain_idx]
    if not isinstance(root, dict):
        return render_template('404.html'), 404
    domain_color = get_domain_color(root.get("name", ""))
    domain_icon = get_domain_icon(root.get("name", ""), domain_idx)
    domain_data = {
        "index": domain_idx,
        "name": root.get("name", ""),
        "color": domain_color,
        "icon": domain_icon,
        "skills": [],
    }
    for si, s in enumerate(root.get("children") or []):
        if not isinstance(s, dict):
            continue
        domain_data["skills"].append(
            {
            "index": si,
            "name": s.get("name", ""),
            "description": s.get("description", ""),
            "responsible": s.get("responsible", ""),
            "level_sticker": s.get("level_sticker", ""),
                "color": get_skill_color(s.get("name", ""), domain_color, si),
                "icon": get_skill_icon(s.get("name", ""), si),
                "actions": [],
            }
        )
    return render_template(
        "domain_view.html", domain=domain_data, current_domain_index=domain_idx, focus_skill=False
    )


@app.route('/domain/<int:domain_idx>/skill/<int:skill_idx>')
def domain_skill_view(domain_idx, skill_idx):
    """Вью навыка: дерево элементов (зависимости от выбранного в сайдбаре)."""
    matrix = get_matrix()
    nodes = (matrix or {}).get("nodes") or []
    if domain_idx < 0 or domain_idx >= _matrix_root_count(matrix):
        return render_template('404.html'), 404
    root = nodes[domain_idx]
    if not isinstance(root, dict):
        return render_template('404.html'), 404
    skills = root.get("children") or []
    if skill_idx < 0 or skill_idx >= len(skills):
        return render_template('404.html'), 404
    skill = skills[skill_idx]
    if not isinstance(skill, dict):
        return render_template('404.html'), 404
    domain_color = get_domain_color(root.get("name", ""))
    domain_icon = get_domain_icon(root.get("name", ""), domain_idx)
    skill_color = get_skill_color(skill.get("name", ""), domain_color, skill_idx)
    skill_icon = get_skill_icon(skill.get("name", ""), skill_idx)
    domain_data = {
        "index": domain_idx,
        "name": root.get("name", ""),
        "color": domain_color,
        "icon": domain_icon,
        "skills": [
            {
            "index": skill_idx,
            "name": skill.get("name", ""),
            "description": skill.get("description", ""),
            "responsible": skill.get("responsible", ""),
            "level_sticker": skill.get("level_sticker", ""),
            "color": skill_color,
            "icon": skill_icon,
                "actions": [],
            }
        ],
    }
    return render_template(
        "domain_view.html",
        domain=domain_data,
        current_domain_index=domain_idx,
        current_skill_index=skill_idx,
        focus_skill=True,
    )


@app.route('/api/matrix')
def api_matrix():
    return jsonify(get_matrix())

@app.route('/api/tree')
def api_tree():
    """Дерево матрицы (корень → листья), уровни определяются автоматически."""
    from core.matrix_schema import annotate_matrix_tree

    meta = get_meta()
    return jsonify(annotate_matrix_tree(get_tree(), meta.get("ui_config") or {}))


@app.route("/api/matrix/change-hints")
def api_matrix_change_hints():
    """leaf_path → автор последней применённой ревизии CR (для наклеек на дереве матрицы)."""
    try:
        matrix = get_matrix()
        live_nodes = matrix.get("nodes") if isinstance(matrix, dict) else None
        with db_session() as session:
            hints = leaf_path_hints_from_applied_changes(
                session,
                live_nodes if isinstance(live_nodes, list) else None,
                limit_crs=80,
            )
        return jsonify({"ok": True, "hints": hints})
    except Exception:
        return jsonify({"ok": True, "hints": {}})


def _leaf_breadcrumb(tree_nodes, path: list) -> str:
    """Строка «Домен → Навык → Действие» для листа по path."""
    if not path:
        return ""
    parts = []
    nodes = tree_nodes
    for i, idx in enumerate(path):
        idx = int(idx)
        if idx < 0 or idx >= len(nodes):
            break
        node = nodes[idx]
        parts.append(node.get("name", ""))
        if i + 1 < len(path) and node.get("children"):
            nodes = node["children"]
    return " → ".join(p for p in parts if p)


_LEAF_VIEW_SOURCE_URL_RE = re.compile(r"https?://[^\s\]\)<>\"']+")


def _coerce_leaf_source_dict(d: Any) -> Dict[str, str]:
    if not isinstance(d, dict):
        return {}
    rid = str(d.get("id") or d.get("literature_id") or "").strip()
    url = str(d.get("url") or d.get("link") or "").strip()
    title = str(d.get("title") or d.get("name") or d.get("text") or "").strip()
    raw = str(d.get("raw") or d.get("value") or "").strip()
    return {
        "literature_id": rid,
        "url": url,
        "title": title or raw,
        "raw": raw or title or url,
    }


def _parse_line_to_source_entry(line: str) -> Optional[Dict[str, str]]:
    line = (line or "").strip()
    if not line:
        return None
    if line.startswith("{") and line.endswith("}"):
        try:
            obj = json.loads(line)
            if isinstance(obj, dict):
                return _coerce_leaf_source_dict(obj)
        except (json.JSONDecodeError, TypeError, ValueError):
            pass
    m = _LEAF_VIEW_SOURCE_URL_RE.search(line)
    url = m.group(0) if m else ""
    rest = line.replace(url, "").strip(" —–-|").strip() if url else line
    title = rest or (url if url else line)
    return _coerce_leaf_source_dict({"url": url, "title": title, "raw": line})


def _parse_leaf_view_source_entries(sources_val: Any) -> List[Dict[str, str]]:
    entries: List[Dict[str, str]] = []
    if sources_val is None:
        return entries
    if isinstance(sources_val, dict):
        d = _coerce_leaf_source_dict(sources_val)
        if d.get("literature_id") or d.get("url") or d.get("title") or d.get("raw"):
            entries.append(d)
        return entries
    if isinstance(sources_val, str):
        for ln in re.split(r"[\n\r]+", sources_val):
            e = _parse_line_to_source_entry(ln)
            if e:
                entries.append(e)
        return entries
    if isinstance(sources_val, list):
        for x in sources_val:
            if isinstance(x, dict):
                d = _coerce_leaf_source_dict(x)
            else:
                d = _parse_line_to_source_entry(str(x))
            if d and (d.get("literature_id") or d.get("url") or d.get("title") or d.get("raw")):
                entries.append(d)
    return entries


def _match_literature_for_source_entry(entry: Dict[str, str], literature: Dict[str, Any]) -> Optional[str]:
    lit = literature or {}
    rid = (entry.get("literature_id") or "").strip()
    if rid and rid in lit:
        return rid
    raw = (entry.get("raw") or "").strip()
    if raw and raw in lit:
        return raw
    url = (entry.get("url") or "").strip()
    if url:
        for r, it in lit.items():
            if (str(it.get("url") or "").strip()) == url:
                return str(r)
    title = (entry.get("title") or "").strip()
    if title:
        tl = title.lower()
        for r, it in lit.items():
            if (str(it.get("title") or "").strip().lower()) == tl:
                return str(r)
    return None


def _resolve_matrix_sources(leaf_view: Optional[Dict[str, Any]], literature: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Слияние leaf_view.sources с каталогом литературы для предпросмотра (как у resource_ids)."""
    lit = literature or {}
    out: List[Dict[str, Any]] = []
    seen: Set[Tuple[Optional[str], str, str]] = set()
    for e in _parse_leaf_view_source_entries((leaf_view or {}).get("sources")):
        rid = _match_literature_for_source_entry(e, lit)
        if rid:
            row = {"id": rid, **(lit.get(rid) or {})}
        else:
            url = (e.get("url") or "").strip()
            title = (e.get("title") or e.get("raw") or url or "Источник").strip()
            row = {
                "title": title,
                "chapter": "",
                "pages": "",
                "url": url,
                "description": "",
                "local_path": "",
            }
        key = (row.get("id"), str(row.get("url") or ""), str(row.get("title") or ""))
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def _literature_linked_leaves(meta: Dict[str, Any]) -> Dict[str, List[Dict[str, str]]]:
    """Для каждого id литературы — листы матрицы (path), где источник задан через шаблон или leaf_view.sources."""
    tree = get_tree()
    templates = meta.get("action_templates", {}) or {}
    literature = meta.get("literature", {}) or {}
    by_rid: Dict[str, List[Dict[str, str]]] = {}
    seen_pairs: Set[Tuple[str, str]] = set()

    def add(rid: str, path_str: str, crumb: str) -> None:
        if rid not in literature:
            return
        k = (rid, path_str)
        if k in seen_pairs:
            return
        seen_pairs.add(k)
        by_rid.setdefault(rid, []).append({"path_str": path_str, "breadcrumb": crumb})

    for leaf in collect_leaves(tree):
        path = leaf.get("path") or []
        path_str = "/".join(str(x) for x in path)
        crumb = _leaf_breadcrumb(tree, path)
        tid = leaf.get("template_id")
        if tid and tid in templates:
            for rid in templates[tid].get("resource_ids", []) or []:
                add(str(rid), path_str, crumb)
        lv = leaf.get("leaf_view") or {}
        for e in _parse_leaf_view_source_entries(lv.get("sources")):
            mrid = _match_literature_for_source_entry(e, literature)
            if mrid:
                add(mrid, path_str, crumb)
    return by_rid


def _matrix_only_source_catalog(meta: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Источники только в leaf_view (без совпадения с каталогом) — виртуальные строки для /api/literature."""
    tree = get_tree()
    literature = meta.get("literature", {}) or {}
    groups: Dict[str, Dict[str, Any]] = {}

    for leaf in collect_leaves(tree):
        path = leaf.get("path") or []
        path_str = "/".join(str(x) for x in path)
        crumb = _leaf_breadcrumb(tree, path)
        lv = leaf.get("leaf_view") or {}
        for e in _parse_leaf_view_source_entries(lv.get("sources")):
            if _match_literature_for_source_entry(e, literature):
                continue
            url = (e.get("url") or "").strip()
            title = (e.get("title") or e.get("raw") or "").strip() or url
            if not title and not url:
                continue
            gk = url or title
            if gk not in groups:
                h = hashlib.md5(gk.encode("utf-8")).hexdigest()[:12]
                display_title = title if (title and title != url) else (url or title)
                groups[gk] = {
                    "id": f"matrix_inline_{h}",
                    "title": display_title or "Источник",
                    "chapter": "",
                    "pages": "",
                    "url": url,
                    "description": "Указано в колонке «Источники» матрицы; отдельной записи в каталоге нет.",
                    "local_path": "",
                    "from_matrix_only": True,
                    "linked_templates": [],
                    "linked_leaves": [],
                }
            gl = groups[gk]["linked_leaves"]
            if not any(x.get("path_str") == path_str for x in gl):
                gl.append({"path_str": path_str, "breadcrumb": crumb})
    return list(groups.values())


@app.route('/api/tree-for-link')
def api_tree_for_link():
    """Дерево для модала привязки: корни → … → листья с template_id (по generic-дереву)."""
    tree = get_tree()
    meta = get_meta()
    templates = meta.get("action_templates", {})
    by_domain: Dict[int, Dict[str, Any]] = {}
    for n in collect_leaves(tree):
        p = n.get("path") or []
        if len(p) < 3:
            continue
        tid = n.get("template_id")
        if not tid or tid not in templates:
            continue
        di, si = int(p[0]), int(p[1])
        anc = get_ancestors(tree, p)
        chain = list(anc) + [n]
        names = [(x.get("name") or "").strip() for x in chain]
        d_name = names[0] if names else ""
        sk_name = names[1] if len(names) > 1 else ""
        leaf_label = names[-1] if names else ""
        if di not in by_domain:
            by_domain[di] = {"name": d_name, "_skills": {}}
        sk_map = by_domain[di]["_skills"]
        if si not in sk_map:
            sk_map[si] = {"name": sk_name, "actions": []}
        sk_map[si]["actions"].append(
            {
                "name": leaf_label,
                "path": p,
                "path_str": "/".join(map(str, p)),
                                "template_id": tid,
            }
        )
    out = []
    for di in sorted(by_domain.keys()):
        d_entry = by_domain[di]
        skills = []
        for si in sorted(d_entry["_skills"].keys()):
            sk = d_entry["_skills"][si]
            if sk["actions"]:
                skills.append({"name": sk["name"], "actions": sk["actions"]})
        if skills:
            out.append({"name": d_entry["name"], "skills": skills})
    return jsonify(out)


@app.route('/api/leaves')
def api_leaves():
    """Список всех листьев с path, template_id, url и иерархией (домен → навык → дерево)."""
    tree = get_tree()
    hierarchy = request.args.get("hierarchy", "").lower() in ("1", "true", "yes")
    leaves = collect_leaves(tree)
    out = []
    for n in leaves:
        p = n.get("path", [])
        item = {
            "path": p,
            "name": n.get("name"),
            "url": path_to_url(p),
            "template_id": n.get("template_id"),
        }
        if hierarchy:
            item["breadcrumb"] = _leaf_breadcrumb(tree, p)
        out.append(item)
    return jsonify(out)

@app.route('/api/leaf-literature')
def api_leaf_literature():
    """Мапа path_str -> список названий привязанной литературы для отображения в матрице."""
    meta = get_meta()
    templates = meta.get("action_templates", {})
    literature = meta.get("literature", {})
    tree = get_tree()
    leaves = collect_leaves(tree)
    out = {}
    for n in leaves:
        p = n.get("path", [])
        tid = n.get("template_id")
        if not tid or tid not in templates:
            continue
        rids = templates[tid].get("resource_ids", [])
        titles = [literature.get(rid, {}).get("title", rid) for rid in rids if rid in literature]
        if titles:
            path_str = "/".join(str(x) for x in p)
            out[path_str] = titles
    return jsonify(out)


@app.route('/api/meta')
def api_meta():
    return jsonify(get_meta())

@app.route('/graph')
def graph():
    return render_template('graph.html')

@app.route('/api/graph-data')
def graph_data():
    meta = get_meta()
    return jsonify(_build_global_graph_from_generic_tree(meta))

# ----- Универсальный маршрут листа (произвольная глубина) -----

@app.route('/leaf/<path:path>')
def leaf_page(path):
    """Страница листа по path (например 0/1/2 или 0/1/2/0)."""
    resolved = resolve_leaf_by_path(path)
    if not resolved:
        abort(404)
    domain, skill, action, parent_action_text = resolved
    path_parts = path.strip("/").split("/")
    di = int(path_parts[0]) if len(path_parts) > 0 else 0
    si = int(path_parts[1]) if len(path_parts) > 1 else 0
    ai = int(path_parts[2]) if len(path_parts) > 2 else 0
    sub_idx = int(path_parts[3]) if len(path_parts) > 3 else None
    domain_color = get_domain_color(domain["name"])
    skill_color = get_skill_color(skill["name"], domain_color, si)
    ctx = {
        "domain": domain,
        "skill": skill,
        "action": action,
        "action_text": action["text"],
        "di": di, "si": si, "ai": ai,
        "domain_color": domain_color,
        "skill_color": skill_color,
        "domain_icon": get_domain_icon(domain["name"], di),
        "skill_icon": get_skill_icon(skill["name"], si),
        "leaf_path": path,
    }
    if sub_idx is not None:
        ctx["sub_idx"] = sub_idx
        ctx["parent_action_text"] = parent_action_text
    return render_template("action_detail.html", **ctx)

@app.route('/api/leaf/<path:path>')
def leaf_api(path):
    """API листа: те же данные, что /api/action или /api/subaction."""
    resolved = resolve_leaf_by_path(path)
    if not resolved:
        return jsonify({"error": "Not found"}), 404
    domain, skill, action, _parent_action_text = resolved
    meta = get_meta()
    template_id = action.get("template_id")
    template = meta.get("action_templates", {}).get(template_id or "", {})
    if template.get("is_parent"):
        return jsonify({"error": "Node is not a leaf"}), 400
    enriched = enrich_action(action, template, meta)
    description = build_description(action, template, domain, skill, meta)
    path_parts = [int(x) for x in path.strip("/").split("/") if x.strip()]
    related = find_related_skills_by_path(path_parts) if len(path_parts) >= 3 else []
    literature_map = meta.get("literature", {}) or {}
    matrix_sources = _resolve_matrix_sources(action.get("leaf_view"), literature_map)
    node_summary_plain = (action.get("description") or "").strip()
    return jsonify({
        "title": action["text"],
        "description": description,
        "node_summary": node_summary_plain,
        "responsible": action.get("responsible") or "",
        "examples": enriched["examples"],
        "tools": enriched["tools"],
        "stack_labels": enriched["stack_labels"],
        "literature": enriched["literature"],
        "matrix_sources": matrix_sources,
        "level_tag": action.get("level_tag"),
        "level_tags": action_level_tags_for_json(action),
        "leaf_view": action.get("leaf_view") or {},
        "review_questions": action.get("review_questions", []),
        "related_skills": related,
        "domain_color": get_domain_color(domain["name"]),
        "skill_color": get_skill_color(skill["name"], get_domain_color(domain["name"]), path_parts[1] if len(path_parts) > 1 else 0),
        "domain_icon": get_domain_icon(domain["name"], path_parts[0] if path_parts else 0),
        "skill_icon": get_skill_icon(skill["name"], path_parts[1] if len(path_parts) > 1 else 0),
        "leaf_path": path,
    })

def find_related_skills_by_path(path_parts: List[int]) -> List[Dict[str, Any]]:
    """По path листа — похожие листья в других ветках (пересечение значимых слов в названии)."""
    if len(path_parts) < 3:
        return []
    tree = get_tree()
    cur = get_node_by_path(tree, path_parts)
    if not cur or cur.get("children"):
        return []
    current_text = (cur.get("name") or "").lower()
    words = set(re.findall(r"\w+", current_text))
    stop_words = {"и", "в", "на", "с", "для", "по", "от", "за", "через", "при", "из", "у", "к", "о", "об"}
    words -= stop_words
    if not words:
        return []
    related: List[Dict[str, Any]] = []
    prefix = list(path_parts)
    for leaf in collect_leaves(tree):
        p = leaf.get("path") or []
        if len(p) < 3 or p == prefix:
            continue
        lt = (leaf.get("name") or "").lower()
        lw = set(re.findall(r"\w+", lt)) - stop_words
        if len(words & lw) < 2:
            continue
        anc = get_ancestors(tree, p)
        dname = anc[0].get("name", "") if anc else ""
        sname = anc[1].get("name", "") if len(anc) > 1 else ""
        ps = "/".join(str(x) for x in p)
        related.append(
            {
                "domain_name": dname,
                "skill_name": sname,
                "action": lt[:60] + ("..." if len(lt) > 60 else ""),
                "url": f"/leaf/{ps}",
            }
        )
        if len(related) >= 5:
            break
    return related

# ----- МАРШРУТЫ ДЛЯ ДЕЙСТВИЙ (обратная совместимость) -----

@app.route('/action/<int:di>/<int:si>/<int:ai>')
def action_page(di, si, ai):
    return redirect(f"/leaf/{di}/{si}/{ai}", code=301)

@app.route('/api/action/<int:di>/<int:si>/<int:ai>')
def action_api(di, si, ai):
    return redirect(f"/api/leaf/{di}/{si}/{ai}", code=307)

# ----- МАРШРУТЫ ДЛЯ ПОДДЕЙСТВИЙ -----

@app.route('/subaction/<int:di>/<int:si>/<int:ai>/<int:sub_idx>')
def subaction_page(di, si, ai, sub_idx):
    return redirect(f"/leaf/{di}/{si}/{ai}/{sub_idx}", code=301)

@app.route('/api/subaction/<int:di>/<int:si>/<int:ai>/<int:sub_idx>')
def subaction_api(di, si, ai, sub_idx):
    return redirect(f"/api/leaf/{di}/{si}/{ai}/{sub_idx}", code=307)

# ----- МАРШРУТЫ ДЛЯ ГРАФОВ ДОМЕНОВ -----

@app.route('/domain-graph/<int:domain_idx>')
def domain_graph(domain_idx):
    data = get_matrix() or {}
    meta = get_meta()
    try:
        if domain_idx < 0 or domain_idx >= _matrix_root_count(data):
            abort(404)
        root = data["nodes"][domain_idx]
        if not isinstance(root, dict):
            abort(404)
        domain = {
            "name": root.get("name", ""),
            "description": root.get("description", ""),
            "responsible": root.get("responsible", ""),
        }
        return render_template(
            "domain_graph.html",
                             domain=domain,
                             domain_idx=domain_idx,
                             current_domain_index=domain_idx,
            ui_config=meta.get("ui_config", {}),
        )
    except (IndexError, KeyError, TypeError) as e:
        print(f"Ошибка при загрузке графа домена: {e}")
        abort(404)

@app.route('/api/domain-graph/<int:domain_idx>')
def domain_graph_data(domain_idx):
    data = get_matrix() or {}
    meta = get_meta()
    try:
        if domain_idx < 0 or domain_idx >= _matrix_root_count(data):
            return jsonify({"error": "Domain not found"}), 404
        
        payload = _build_domain_graph_from_generic_tree(domain_idx, meta)
        if not payload:
            return jsonify({"error": "Domain not found"}), 404
        return jsonify(payload)
    except Exception as e:
        print(f"Ошибка при создании графа домена: {e}")
        return jsonify({"error": str(e)}), 500

# ----- МАРШРУТ ДЛЯ ЭКСПОРТА -----

@app.route('/export')
def export():
    return render_template('export.html')


@app.route("/api/export/unified-table")
def api_export_unified_table():
    """Строки экспорта: порядок и подписи колонок как в импортированной matrix_column_schema (unified relational)."""
    from core.excel_unified_export import build_unified_export_table

    matrix = get_matrix()
    nodes = list(matrix.get("nodes") or [])
    raw = (request.args.get("domains") or "").strip()
    if raw:
        idxs = _parse_export_domain_idxs(raw)
        nodes = [nodes[i] for i in idxs if 0 <= i < len(nodes)]
    meta = get_meta()
    headers, rows = build_unified_export_table(
        [],
        meta.get("ui_config"),
        nodes=(nodes or None),
        include_header_tags=False,
    )
    return jsonify(
        {
            "ok": True,
            "headers": headers,
            "rows": rows,
            "sheet": "Unified_Relational_Span",
        }
    )


@app.route("/api/export/unified.xlsx")
def api_export_unified_xlsx():
    """Скачивание XLSX: лист Unified_Relational_Span + опционально Литература."""
    try:
        from io import BytesIO

        from openpyxl import Workbook
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl не установлен (pip install openpyxl)"}), 500

    from core.excel_unified_export import build_unified_export_table

    matrix = get_matrix()
    nodes = list(matrix.get("nodes") or [])
    raw = (request.args.get("domains") or "").strip()
    if raw:
        idxs = _parse_export_domain_idxs(raw)
        nodes = [nodes[i] for i in idxs if 0 <= i < len(nodes)]
    meta = get_meta()
    # Строка заголовков с тегами (item)/(leaf_view)/… — та же форма, что ожидает импорт unified xlsx.
    headers, rows = build_unified_export_table(
        [],
        meta.get("ui_config"),
        nodes=(nodes or None),
        include_header_tags=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Unified_Relational_Span"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    lit = meta.get("literature") or {}
    templates = meta.get("action_templates") or {}
    if isinstance(lit, dict) and lit:
        template_to_leaves: Dict[str, List[str]] = {}
        if isinstance(templates, dict):
            for tid, tpl in templates.items():
                if not isinstance(tpl, dict):
                    continue
                for rid in tpl.get("resource_ids") or []:
                    template_to_leaves.setdefault(str(rid), []).append(str(tpl.get("name") or tid))

        ws2 = wb.create_sheet("Литература")
        ws2.append(["Название", "Глава / раздел", "Страницы", "URL", "Локальный файл", "Привязка к компетенциям"])
        for rid, item in lit.items():
            if not isinstance(item, dict):
                continue
            ws2.append(
                [
                    str(item.get("title") or rid),
                    str(item.get("chapter") or ""),
                    str(item.get("pages") or ""),
                    str(item.get("url") or ""),
                    str(item.get("local_path") or item.get("file_path") or ""),
                    "; ".join(template_to_leaves.get(str(rid), [])),
                ]
            )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"matrix_unified_{datetime.now(timezone.utc).strftime('%Y-%m-%d')}.xlsx",
    )


# ----- СХЕМА И ВАЛИДАЦИЯ -----

@app.route('/api/schema')
def api_schema():
    """Информация о схеме источника (для догрузки и валидации)."""
    return jsonify({"ok": True, **get_schema_info()})


@app.route('/api/validate', methods=["POST"])
def api_validate():
    """Валидация структуры источника. Body: {domains: [...], action_templates: {...}, ...}."""
    data = request.get_json() or {}
    vr = validate_source(data)
    return jsonify({"ok": vr.ok, **vr.to_dict()})


# ----- НАСТРОЙКИ: бэкапы и восстановление -----

@app.route('/import')
def import_page():
    """Редирект в административный импорт."""
    return redirect(url_for("admin_import_page"))


@app.route('/admin/import', methods=["GET"])
def admin_import_page():
    """Импорт структуры/данных из Excel только для admin."""
    actor, role = _extract_actor_role()
    if role != "admin":
        return redirect(url_for("login", next=request.path))
    return render_template('import.html', actor=actor, role=role)


@app.route('/constructor')
def constructor_page():
    """Конструктор изменений матрицы для пользователей."""
    return render_template('constructor.html')


@app.route('/about')
def about_page():
    """Раздел «О приложении»."""
    return render_template('about.html')


@app.route('/settings')
def settings_page():
    return render_template('settings.html')


@app.route('/changes')
def changes_page():
    return render_template('changes.html')


@app.route('/admin/users')
def admin_users_page():
    actor, role = _extract_actor_role()
    if role != "admin":
        return redirect(url_for("login", next=request.path))
    return render_template('admin_users.html', actor=actor, role=role)


@app.route('/api/admin/users', methods=["GET"])
def api_admin_users_list():
    actor, admin_err = _require_admin()
    if admin_err:
        return admin_err
    _ensure_db_schema()
    with db_session() as session:
        rows = session.execute(select(User).order_by(User.id.asc())).scalars().all()
        items = [
            {
                "id": u.id,
                "username": u.username,
                "role": u.role,
                "full_name": u.full_name or "",
                "email": u.email or "",
                "must_change_password": bool(u.must_change_password),
                "is_active": bool(u.is_active),
                "created_at": u.created_at.isoformat() if u.created_at else None,
                "updated_at": u.updated_at.isoformat() if u.updated_at else None,
                "edit_mode": (
                    "none"
                    if _is_e2e_admin_username(u.username)
                    else ("email_only" if _is_system_admin_username(u.username) else "full")
                ),
                "requested_by": actor,
            }
            for u in rows
        ]
    return jsonify({
        "ok": True,
        "items": items,
    })


@app.route('/api/admin/users', methods=["POST"])
def api_admin_users_create():
    data = request.get_json(silent=True) or {}
    _, admin_err = _require_admin()
    if admin_err:
        return admin_err
    username = (data.get("username") or "").strip()
    role = (data.get("role") or "user").strip().lower()
    full_name = (data.get("full_name") or "").strip()
    email = (data.get("email") or "").strip()
    temp_password = (data.get("temp_password") or "").strip()
    if not username:
        return jsonify({"ok": False, "error": "username required"}), 400
    if role not in ("user", "admin"):
        return jsonify({"ok": False, "error": "role must be user or admin"}), 400
    if not _is_valid_email_mask(email):
        return jsonify({"ok": False, "error": "invalid email"}), 400
    if len(temp_password) < 10:
        return jsonify({"ok": False, "error": "temp_password must be at least 10 characters"}), 400
    _ensure_db_schema()
    with db_session() as session:
        exists = session.execute(select(User).where(User.username == username)).scalars().first()
        if exists:
            return jsonify({"ok": False, "error": "username already exists"}), 409
        if _is_e2e_admin_username(username):
            resolved_full_name = full_name or "E2E Admin"
        elif _is_system_admin_username(username):
            resolved_full_name = ADMIN_DISPLAY_NAME
        else:
            resolved_full_name = _default_full_name_for_role(role)
        user = User(
            username=username,
            role=role,
            full_name=resolved_full_name,
            email=email,
            password_hash=generate_password_hash(temp_password),
            must_change_password=True,
            is_active=True,
        )
        session.add(user)
        session.flush()
        created_id = user.id
    return jsonify({"ok": True, "id": created_id, "must_change_password": True})


@app.route('/api/admin/users/<int:user_id>', methods=["PATCH"])
def api_admin_users_update(user_id: int):
    data = request.get_json(silent=True) or {}
    _, admin_err = _require_admin()
    if admin_err:
        return admin_err
    new_username = data.get("username")
    new_role = data.get("role")
    new_full_name = data.get("full_name")
    new_email = data.get("email")
    new_active = data.get("is_active")
    reset_temp_password = data.get("reset_temp_password")
    _ensure_db_schema()
    with db_session() as session:
        user = session.get(User, user_id)
        if not user:
            return jsonify({"ok": False, "error": "user not found"}), 404
        if _is_e2e_admin_username(user.username):
            return jsonify({"ok": False, "error": "e2e_admin can only be deleted"}), 403
        if _is_system_admin_username(user.username):
            has_non_email_update = any(
                x is not None
                for x in (new_username, new_role, new_full_name, new_active, reset_temp_password)
            )
            if has_non_email_update:
                return jsonify({"ok": False, "error": "System Administrator allows email update only"}), 403
        info_updated = False
        if new_username is not None:
            username = (new_username or "").strip()
            if not username:
                return jsonify({"ok": False, "error": "username cannot be empty"}), 400
            exists = session.execute(
                select(User).where(User.username == username, User.id != user_id)
            ).scalars().first()
            if exists:
                return jsonify({"ok": False, "error": "username already exists"}), 409
            if user.username != username:
                user.username = username
                info_updated = True
        if new_role is not None:
            role = (new_role or "").strip().lower()
            if role not in ("user", "admin"):
                return jsonify({"ok": False, "error": "role must be user or admin"}), 400
            if user.role == "admin" and role != "admin":
                admins = session.execute(select(User).where(User.role == "admin", User.is_active == True)).scalars().all()
                if len(admins) <= 1:
                    return jsonify({"ok": False, "error": "cannot demote last active admin"}), 409
            if user.role != role:
                user.role = role
                info_updated = True
                if (
                    not _is_system_admin_username(user.username)
                    and not _is_e2e_admin_username(user.username)
                    and (new_full_name is None or (new_full_name or "").strip() == "")
                ):
                    user.full_name = _default_full_name_for_role(role)
        if new_full_name is not None:
            full_name = (new_full_name or "").strip()
            if not full_name:
                return jsonify({"ok": False, "error": "full_name cannot be empty"}), 400
            if user.full_name != full_name:
                user.full_name = full_name
                info_updated = True
        if new_email is not None:
            email = (new_email or "").strip()
            if not _is_valid_email_mask(email):
                return jsonify({"ok": False, "error": "invalid email"}), 400
            if user.email != email:
                user.email = email
                info_updated = True
        if new_active is not None:
            active = bool(new_active)
            if user.role == "admin" and user.is_active and not active:
                admins = session.execute(select(User).where(User.role == "admin", User.is_active == True)).scalars().all()
                if len(admins) <= 1:
                    return jsonify({"ok": False, "error": "cannot deactivate last active admin"}), 409
            user.is_active = active
        if reset_temp_password is not None:
            temp_password = (reset_temp_password or "").strip()
            if len(temp_password) < 10:
                return jsonify({"ok": False, "error": "reset_temp_password must be at least 10 characters"}), 400
            user.password_hash = generate_password_hash(temp_password)
            user.must_change_password = True
        if info_updated:
            user.updated_at = _utcnow()
    return jsonify({"ok": True})


@app.route('/api/admin/users/<int:user_id>', methods=["DELETE"])
def api_admin_users_delete(user_id: int):
    _, admin_err = _require_admin()
    if admin_err:
        return admin_err
    _ensure_db_schema()
    with db_session() as session:
        user = session.get(User, user_id)
        if not user:
            return jsonify({"ok": False, "error": "user not found"}), 404
        if user.role == "admin":
            admins = session.execute(select(User).where(User.role == "admin", User.is_active == True)).scalars().all()
            if len(admins) <= 1:
                return jsonify({"ok": False, "error": "cannot delete last admin"}), 409
        session.delete(user)
    return jsonify({"ok": True})


@app.route('/admin/sql-console', methods=["GET"])
def admin_sql_console_page():
    actor, role = _extract_actor_role()
    if role != "admin":
        return redirect(url_for("login", next=request.path))
    return render_template('admin_sql_console.html', actor=actor, role=role)


@app.route('/admin/tree-editor', methods=["GET"])
def admin_tree_editor_page():
    actor, role = _extract_actor_role()
    if role != "admin":
        return redirect(url_for("login", next=request.path))
    return render_template('admin_tree_editor.html', actor=actor, role=role)


@app.route('/admin/notifications', methods=["GET"])
def admin_notifications_page():
    actor, role = _extract_actor_role()
    if role != "admin":
        return redirect(url_for("login", next=request.path))
    return render_template('admin_notifications.html', actor=actor, role=role)


@app.route('/admin/presence', methods=["GET"])
def admin_presence_page():
    actor, role = _extract_actor_role()
    if role != "admin":
        return redirect(url_for("login", next=request.path))
    return render_template('admin_presence.html', actor=actor, role=role)


@app.route('/api/admin/tree-editor/data', methods=["GET"])
def api_admin_tree_editor_data():
    _, admin_err = _require_admin()
    if admin_err:
        return admin_err
    _ensure_db_schema()
    with db_session() as session:
        unified = load_unified_from_db(session, literature=load_literature_map())
    ui = unified.get("ui_config") or {}
    constructor_levels, _constructor_extra = build_constructor_levels(ui)
    return jsonify(
        {
            "ok": True,
            "nodes": unified.get("nodes") or [],
            "action_templates": unified.get("action_templates") or {},
            "literature": unified.get("literature") or {},
            "constructor_levels": constructor_levels,
        }
    )


@app.route('/api/constructor/meta', methods=["GET"])
def api_constructor_meta():
    auth, auth_err = _require_authenticated()
    if auth_err:
        return auth_err
    _ensure_db_schema()
    with db_session() as session:
        unified = load_unified_from_db(session, literature=load_literature_map())
    templates = unified.get("action_templates") or {}
    template_ids = sorted([str(k) for k in templates.keys() if str(k).strip()])
    template_list = [
        {"id": str(tid), "name": str((tpl or {}).get("name") or tid).strip() or str(tid)}
        for tid, tpl in sorted(templates.items(), key=lambda x: str(x[0]))
        if str(tid).strip()
    ]
    ui = unified.get("ui_config") or {}
    matrix_levels = merge_matrix_levels(ui)
    constructor_levels, constructor_extra_leaf_steps = build_constructor_levels(ui)
    mcs_eff = effective_matrix_column_schema(ui)
    mcs = schema_entries_for_ui(mcs_eff, ui)
    unified_cols = mcs
    return jsonify(
        {
            "ok": True,
            "requested_by": auth["actor"],
            "nodes": unified.get("nodes") or [],
            "template_ids": template_ids,
            "templates": template_list,
            "matrix_levels": matrix_levels,
            "matrix_column_schema": mcs,
            "unified_column_schema": unified_cols,
            "constructor_leaf_step_title": str(ui.get("constructor_leaf_step_title") or "").strip(),
            "constructor_levels": constructor_levels,
            "constructor_tower": {"extra_leaf_steps": constructor_extra_leaf_steps},
            "constructor_value_lists": {
                "sticker_grades": list(STICKER_GRADES),
                "templates": template_list,
                "template_ids": template_ids,
            },
            "sticker_grades": list(STICKER_GRADES),
            "skill_sticker_tag": TAG_SKILL_STICKER,
            "updated_at": _utcnow().isoformat(),
        }
    )


@app.route('/api/constructor/preview', methods=["POST"])
def api_constructor_preview():
    auth, auth_err = _require_authenticated()
    if auth_err:
        return auth_err
    data = request.get_json(silent=True) or {}
    payload = data.get("payload") or {}
    merge_mode = (data.get("merge_mode") or "append").strip()
    target_domain = data.get("target_domain")
    target_skill = data.get("target_skill")
    if not isinstance(payload, dict):
        return jsonify({"ok": False, "error": "payload must be object"}), 400

    _ensure_db_schema()
    with db_session() as session:
        current = load_unified_from_db(session, literature=load_literature_map())
        merged = merge_upload_into_source(
            current,
            payload,
            merge_mode=merge_mode,
            target_domain=target_domain,
            target_skill=target_skill,
        )
        revision_payload = build_revision_payload(
            base_snapshot=current,
            upload_payload=payload,
            proposed_snapshot=merged,
            merge_mode=merge_mode,
            target_domain=target_domain,
            target_skill=target_skill,
        )
    return jsonify(
        {
            "ok": True,
            "requested_by": auth["actor"],
            "merge_mode": merge_mode,
            "target_domain": target_domain,
            "target_skill": target_skill,
            "diff": revision_payload.get("structural_diff") or {},
            "json_patch_ops": len(revision_payload.get("json_patch") or []),
            "upsert_plan": revision_payload.get("upsert_plan") or {},
            "payload_preview": payload,
        }
    )


@app.route('/api/admin/tree-editor/preview', methods=["POST"])
def api_admin_tree_editor_preview():
    payload = request.get_json(silent=True) or {}
    actor, admin_err = _require_admin(payload)
    if admin_err:
        return admin_err
    edited_nodes = payload.get("nodes")
    if not isinstance(edited_nodes, list):
        return jsonify({"ok": False, "error": "nodes array is required"}), 400
    _ensure_db_schema()
    with db_session() as session:
        current = load_unified_from_db(session, literature=load_literature_map())
    clean_nodes = strip_transient_node_fields(deepcopy(edited_nodes))
    proposed = dict(current)
    proposed["nodes"] = clean_nodes
    proposed["domains"] = []
    warnings = _build_tree_edit_warnings(current, clean_nodes)
    revision_payload = build_revision_payload(
        base_snapshot=current,
        upload_payload={"nodes": clean_nodes},
        proposed_snapshot=proposed,
        merge_mode="replace_all",
    )
    return jsonify(
        {
            "ok": True,
            "requested_by": actor,
            "warnings": warnings,
            "diff": revision_payload.get("structural_diff"),
            "json_patch_ops": len(revision_payload.get("json_patch") or []),
            "upsert_plan": revision_payload.get("upsert_plan") or {},
        }
    )


@app.route('/api/admin/tree-editor/submit', methods=["POST"])
def api_admin_tree_editor_submit():
    payload = request.get_json(silent=True) or {}
    actor, admin_err = _require_admin(payload)
    if admin_err:
        return admin_err
    edited_nodes = payload.get("nodes")
    if not isinstance(edited_nodes, list):
        return jsonify({"ok": False, "error": "nodes array is required"}), 400
    title = (payload.get("title") or "").strip() or "Admin tree edit"
    confirm_rel = bool(payload.get("confirm_relations"))
    _ensure_db_schema()
    with db_session() as session:
        current = load_unified_from_db(session, literature=load_literature_map())
    clean_nodes = strip_transient_node_fields(deepcopy(edited_nodes))
    warnings = _build_tree_edit_warnings(current, clean_nodes)
    if warnings["removed_template_count"] > 0 and not confirm_rel:
        return jsonify(
            {
                "ok": False,
                "error": "relations_confirmation_required",
                "warnings": warnings,
                "message": "Tree edit removes template bindings. Confirm relation override.",
            }
        ), 409
    proposed = dict(current)
    proposed["nodes"] = clean_nodes
    proposed["domains"] = []
    revision_payload = build_revision_payload(
        base_snapshot=current,
        upload_payload={"nodes": clean_nodes},
        proposed_snapshot=proposed,
        merge_mode="replace_all",
    )
    with db_session() as session:
        cr = create_change_request(
            session=session,
            title=title,
            merge_mode="replace_all",
            payload=revision_payload,
            created_by=actor,
        )
        approval_set_status(session, cr.id, "submitted", actor=actor, comment="Admin tree edit submitted")
        _notify_cr_submitted(session, cr)
        change_id = cr.id
    return jsonify(
        {
            "ok": True,
            "change_id": change_id,
            "status": "submitted",
            "warnings": warnings,
            "message": "Changes submitted for approval with diff.",
        }
    )


@app.route('/api/admin/db/objects', methods=["GET"])
def api_admin_db_objects():
    actor, admin_err = _require_admin()
    if admin_err:
        return admin_err
    _ensure_db_schema()
    with db_session() as session:
        tables = session.execute(
            text(
                """
                SELECT table_schema AS schema, table_name AS name, table_type
                FROM information_schema.tables
                WHERE table_schema NOT IN ('pg_catalog', 'information_schema')
                ORDER BY table_schema, table_name
                """
            )
        ).mappings().all()
        sequences = session.execute(
            text(
                """
                SELECT sequence_schema AS schema, sequence_name AS name
                FROM information_schema.sequences
                WHERE sequence_schema NOT IN ('pg_catalog', 'information_schema')
                ORDER BY sequence_schema, sequence_name
                """
            )
        ).mappings().all()
        functions = session.execute(
            text(
                """
                SELECT n.nspname AS schema, p.proname AS name, pg_get_function_identity_arguments(p.oid) AS args
                FROM pg_proc p
                JOIN pg_namespace n ON n.oid = p.pronamespace
                WHERE n.nspname NOT IN ('pg_catalog', 'information_schema')
                ORDER BY n.nspname, p.proname
                """
            )
        ).mappings().all()
    items = []
    for row in tables:
        table_type = row["table_type"]
        if table_type == "BASE TABLE":
            obj_type = "table"
        elif table_type == "VIEW":
            obj_type = "view"
        else:
            obj_type = "table_like"
        items.append(
            {
                "type": obj_type,
                "schema": row["schema"],
                "name": row["name"],
                "qualified_name": f"{row['schema']}.{row['name']}",
            }
        )
    for row in sequences:
        items.append(
            {
                "type": "sequence",
                "schema": row["schema"],
                "name": row["name"],
                "qualified_name": f"{row['schema']}.{row['name']}",
            }
        )
    for row in functions:
        args = row["args"] or ""
        items.append(
            {
                "type": "function",
                "schema": row["schema"],
                "name": row["name"],
                "args": args,
                "qualified_name": f"{row['schema']}.{row['name']}({args})",
            }
        )
    return jsonify({"ok": True, "requested_by": actor, "items": items})


@app.route('/api/admin/db/object/details', methods=["POST"])
def api_admin_db_object_details():
    payload = request.get_json(silent=True) or {}
    _, admin_err = _require_admin(payload)
    if admin_err:
        return admin_err
    schema = (payload.get("schema") or "").strip()
    name = (payload.get("name") or "").strip()
    obj_type = (payload.get("type") or "table").strip().lower()
    row_limit = int(payload.get("row_limit") or 25)
    row_limit = min(max(row_limit, 1), 200)
    if not schema or not name:
        return jsonify({"ok": False, "error": "schema and name are required"}), 400
    try:
        _quote_ident(schema)
        _quote_ident(name)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400

    _ensure_db_schema()
    ddl = ""
    columns = []
    rows = []
    try:
        with db_session() as session:
            if obj_type == "table":
                ddl = _build_table_ddl(session, schema, name)
                qname = _qualified_ident(schema, name)
                rows_result = session.execute(text(f"SELECT * FROM {qname} LIMIT {row_limit}"))
                columns = list(rows_result.keys())
                rows = [list(r) for r in rows_result.fetchall()]
            elif obj_type == "view":
                ddl_row = session.execute(
                    text(
                        """
                        SELECT pg_get_viewdef((quote_ident(:schema) || '.' || quote_ident(:name))::regclass, true) AS ddl
                        """
                    ),
                    {"schema": schema, "name": name},
                ).mappings().first()
                ddl = f"CREATE VIEW {_qualified_ident(schema, name)} AS\n{(ddl_row or {}).get('ddl') or '-- view not found'}"
                qname = _qualified_ident(schema, name)
                rows_result = session.execute(text(f"SELECT * FROM {qname} LIMIT {row_limit}"))
                columns = list(rows_result.keys())
                rows = [list(r) for r in rows_result.fetchall()]
            elif obj_type == "sequence":
                ddl = f"-- Sequence: {_qualified_ident(schema, name)}"
                qname = _qualified_ident(schema, name)
                rows_result = session.execute(text(f"SELECT * FROM {qname}"))
                columns = list(rows_result.keys())
                rows = [list(r) for r in rows_result.fetchall()]
            elif obj_type == "function":
                ddl = f"-- Function scanner entry: {schema}.{name}"
            else:
                return jsonify({"ok": False, "error": f"Unsupported object type: {obj_type}"}), 400
    except Exception as exc:
        return jsonify({"ok": False, "error": _format_db_error(exc)}), 400

    return jsonify(
        {
            "ok": True,
            "object": {"type": obj_type, "schema": schema, "name": name},
            "ddl": ddl,
            "columns": columns,
            "rows": rows,
        }
    )


@app.route('/api/admin/db/execute', methods=["POST"])
def api_admin_db_execute():
    payload = request.get_json(silent=True) or {}
    _, admin_err = _require_admin(payload)
    if admin_err:
        return admin_err
    sql = (payload.get("sql") or "").strip()
    if not sql:
        return jsonify({"ok": False, "error": "sql is required"}), 400
    _ensure_db_schema()
    try:
        with db_session() as session:
            result = session.execute(text(sql))
            if result.returns_rows:
                columns = list(result.keys())
                rows = [list(r) for r in result.fetchall()]
                return jsonify(
                    {
                        "ok": True,
                        "returns_rows": True,
                        "columns": columns,
                        "rows": rows,
                        "row_count": len(rows),
                    }
                )
            return jsonify(
                {
                    "ok": True,
                    "returns_rows": False,
                    "row_count": int(result.rowcount or 0),
                }
            )
    except Exception as exc:
        return jsonify({"ok": False, "error": _format_db_error(exc)}), 400


@app.route('/api/admin/db/diagram', methods=["GET"])
def api_admin_db_diagram():
    _, admin_err = _require_admin()
    if admin_err:
        return admin_err
    _ensure_db_schema()
    with db_session() as session:
        tables = session.execute(
            text(
                """
                SELECT table_schema AS schema, table_name AS name
                FROM information_schema.tables
                WHERE table_schema = 'public' AND table_type = 'BASE TABLE'
                ORDER BY table_name
                """
            )
        ).mappings().all()
        edges = session.execute(
            text(
                """
                SELECT
                    tc.table_name AS from_table,
                    ccu.table_name AS to_table,
                    tc.constraint_name AS constraint_name
                FROM information_schema.table_constraints tc
                JOIN information_schema.key_column_usage kcu
                    ON tc.constraint_name = kcu.constraint_name
                    AND tc.table_schema = kcu.table_schema
                JOIN information_schema.constraint_column_usage ccu
                    ON ccu.constraint_name = tc.constraint_name
                    AND ccu.table_schema = tc.table_schema
                WHERE tc.constraint_type = 'FOREIGN KEY'
                  AND tc.table_schema = 'public'
                ORDER BY tc.table_name, ccu.table_name
                """
            )
        ).mappings().all()
    table_names = [row["name"] for row in tables]
    mermaid_lines = ["erDiagram"]
    for t in table_names:
        mermaid_lines.append(f"  {t} {{")
        mermaid_lines.append("    int id")
        mermaid_lines.append("  }")
    for edge in edges:
        left = edge["to_table"]
        right = edge["from_table"]
        label = (edge["constraint_name"] or "fk").replace('"', "")
        mermaid_lines.append(f"  {left} ||--o{{ {right} : \"{label}\"")
    return jsonify(
        {
            "ok": True,
            "tables": table_names,
            "edges": [dict(e) for e in edges],
            "mermaid": "\n".join(mermaid_lines),
        }
    )


@app.route('/api/admin/notifications', methods=["GET"])
def api_admin_notifications_list():
    _, admin_err = _require_admin()
    if admin_err:
        return admin_err
    status_filter = (request.args.get("status") or "").strip().lower()
    q_filter = (request.args.get("q") or "").strip()
    only_failed = (request.args.get("only_failed") or "").strip().lower() in ("1", "true", "yes")
    try:
        limit_value = int((request.args.get("limit") or "100").strip())
    except ValueError:
        limit_value = 100
    limit = max(1, min(500, limit_value))
    _ensure_db_schema()
    with db_session() as session:
        stmt = select(NotificationLog).order_by(NotificationLog.id.desc())
        if only_failed and not status_filter:
            stmt = stmt.where(NotificationLog.status.in_(["failed", "skipped"]))
        if status_filter:
            stmt = stmt.where(NotificationLog.status == status_filter)
        if q_filter:
            like = f"%{q_filter}%"
            stmt = stmt.where(
                or_(
                    NotificationLog.event_type.ilike(like),
                    NotificationLog.subject.ilike(like),
                    NotificationLog.error.ilike(like),
                )
            )
        rows = session.execute(stmt.limit(limit)).scalars().all()
        if q_filter:
            ql = q_filter.lower()
            rows = [
                n
                for n in rows
                if ql in ",".join(n.recipients or []).lower() or ql in str(n.context or {}).lower() or ql in (n.created_by or "").lower()
            ]
        items = [
            {
                "id": n.id,
                "event_type": n.event_type,
                "status": n.status,
                "subject": n.subject,
                "recipients": n.recipients or [],
                "error": n.error,
                "attempts": n.attempts,
                "created_by": n.created_by,
                "created_at": n.created_at.isoformat() if n.created_at else None,
                "last_attempt_at": n.last_attempt_at.isoformat() if n.last_attempt_at else None,
                "sent_at": n.sent_at.isoformat() if n.sent_at else None,
                "context": n.context or {},
            }
            for n in rows
        ]
    return jsonify({"ok": True, "items": items})


@app.route('/api/admin/notifications/<int:notification_id>/retry', methods=["POST"])
def api_admin_notifications_retry(notification_id: int):
    actor, admin_err = _require_admin()
    if admin_err:
        return admin_err
    _ensure_db_schema()
    with db_session() as session:
        row = session.get(NotificationLog, notification_id)
        if not row:
            return jsonify({"ok": False, "error": "Notification not found"}), 404
        ok, err = _deliver_notification_log(row)
        row.context = {**(row.context or {}), "last_retry_by": actor}
        out_status = row.status
    return jsonify({"ok": ok, "status": out_status, "error": err})


@app.route('/api/admin/presence', methods=["GET"])
def api_admin_presence():
    _, admin_err = _require_admin()
    if admin_err:
        return admin_err
    username_filter = (request.args.get("username") or "").strip().lower()
    status_filter = (request.args.get("status") or "").strip().lower()
    try:
        limit = max(1, min(500, int((request.args.get("limit") or "200").strip())))
    except ValueError:
        limit = 200

    rows, summary = _presence_rows_summary(username_filter, status_filter, limit)

    return jsonify({"ok": True, "summary": summary, "items": rows})


def _presence_rows_summary(username_filter: str, status_filter: str, limit: int) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    now = _utcnow()
    week_start = now - timedelta(days=7)
    _ensure_db_schema()
    with db_session() as db:
        users = db.execute(select(User).where(User.is_active == True).order_by(User.username.asc())).scalars().all()
        sessions = db.execute(
            select(UserPresenceSession).where(
                or_(
                    UserPresenceSession.logout_at.is_(None),
                    UserPresenceSession.last_seen_at >= week_start,
                    UserPresenceSession.login_at >= week_start,
                )
            )
        ).scalars().all()

    by_username: Dict[str, List[UserPresenceSession]] = {}
    for s in sessions:
        uname = (s.username or "").strip()
        if not uname:
            continue
        by_username.setdefault(uname, []).append(s)

    rows: List[Dict[str, Any]] = []
    for u in users:
        uname = (u.username or "").strip()
        user_sessions = by_username.get(uname, [])
        active = [s for s in user_sessions if s.logout_at is None]
        latest_seen = max((s.last_seen_at for s in user_sessions), default=None)
        active_latest_seen = max((s.last_seen_at for s in active), default=None)
        if active_latest_seen is not None:
            delta = (now - active_latest_seen).total_seconds()
            if delta <= PRESENCE_ONLINE_SECONDS:
                status = "online"
            elif delta <= PRESENCE_AWAY_SECONDS:
                status = "away"
            else:
                status = "offline"
        else:
            status = "offline"

        current_session_seconds = sum(max(0, int((now - s.login_at).total_seconds())) for s in active)
        week_total_seconds = 0
        week_sessions_count = 0
        last_session_seconds: Optional[int] = None
        ended_sorted = sorted([s for s in user_sessions if s.logout_at is not None], key=lambda x: x.logout_at or x.last_seen_at, reverse=True)
        if ended_sorted:
            last_end = ended_sorted[0].logout_at or ended_sorted[0].last_seen_at
            last_session_seconds = max(0, int((last_end - ended_sorted[0].login_at).total_seconds()))

        for s in user_sessions:
            end = s.logout_at or now
            start = s.login_at
            overlap_start = max(start, week_start)
            overlap_end = min(end, now)
            if overlap_end > overlap_start:
                week_total_seconds += int((overlap_end - overlap_start).total_seconds())
                week_sessions_count += 1

        row = {
            "user_id": u.id,
            "username": uname,
            "full_name": u.full_name or "",
            "email": u.email or "",
            "status": status,
            "last_seen_at": latest_seen.isoformat() if latest_seen else None,
            "active_sessions": len(active),
            "current_session_seconds": current_session_seconds,
            "last_session_seconds": last_session_seconds,
            "week_total_seconds": week_total_seconds,
            "week_sessions_count": week_sessions_count,
        }
        if username_filter and username_filter not in uname.lower() and username_filter not in (u.full_name or "").lower():
            continue
        if status_filter and status_filter != "all" and row["status"] != status_filter:
            continue
        rows.append(row)

    rows = sorted(
        rows,
        key=lambda x: (0 if x["status"] == "online" else (1 if x["status"] == "away" else 2), -(x["week_total_seconds"] or 0), x["username"]),
    )[:limit]

    summary = {
        "total_users": len(rows),
        "online_users": len([r for r in rows if r["status"] == "online"]),
        "away_users": len([r for r in rows if r["status"] == "away"]),
        "offline_users": len([r for r in rows if r["status"] == "offline"]),
        "week_total_seconds": sum(r["week_total_seconds"] for r in rows),
    }
    return rows, summary


@app.route('/api/admin/presence/<username>/sessions', methods=["GET"])
def api_admin_presence_sessions(username: str):
    _, admin_err = _require_admin()
    if admin_err:
        return admin_err
    uname = (username or "").strip()
    if not uname:
        return jsonify({"ok": False, "error": "username is required"}), 400
    try:
        limit = max(1, min(200, int((request.args.get("limit") or "20").strip())))
    except ValueError:
        limit = 20
    try:
        offset = max(0, int((request.args.get("offset") or "0").strip()))
    except ValueError:
        offset = 0
    from_str = (request.args.get("from") or "").strip()
    to_str = (request.args.get("to") or "").strip()
    sort_by = (request.args.get("sort_by") or "login_at").strip().lower()
    sort_order = (request.args.get("sort_order") or "desc").strip().lower()
    from_dt: Optional[datetime] = None
    to_dt: Optional[datetime] = None
    if from_str:
        try:
            from_dt = datetime.strptime(from_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            return jsonify({"ok": False, "error": "invalid 'from' date format, expected YYYY-MM-DD"}), 400
    if to_str:
        try:
            to_dt = datetime.strptime(to_str, "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
        except ValueError:
            return jsonify({"ok": False, "error": "invalid 'to' date format, expected YYYY-MM-DD"}), 400
    _ensure_db_schema()
    with db_session() as db:
        filters = [UserPresenceSession.username == uname]
        if from_dt is not None:
            filters.append(UserPresenceSession.login_at >= from_dt)
        if to_dt is not None:
            filters.append(UserPresenceSession.login_at < to_dt)
        total = int(db.execute(select(func.count(UserPresenceSession.id)).where(*filters)).scalar_one() or 0)
        duration_seconds_expr = func.extract(
            "epoch",
            func.coalesce(UserPresenceSession.logout_at, func.now()) - UserPresenceSession.login_at,
        )
        sortable = {
            "login_at": UserPresenceSession.login_at,
            "last_seen_at": UserPresenceSession.last_seen_at,
            "logout_at": UserPresenceSession.logout_at,
            "duration_seconds": duration_seconds_expr,
            "id": UserPresenceSession.id,
        }
        sort_expr = sortable.get(sort_by, UserPresenceSession.login_at)
        order_expr = asc(sort_expr) if sort_order == "asc" else desc(sort_expr)
        rows = db.execute(
            select(UserPresenceSession)
            .where(*filters)
            .order_by(order_expr, UserPresenceSession.id.desc())
            .offset(offset)
            .limit(limit)
        ).scalars().all()
    now = _utcnow()
    items = []
    for s in rows:
        end = s.logout_at or now
        duration_seconds = max(0, int((end - s.login_at).total_seconds()))
        items.append(
            {
                "id": s.id,
                "username": s.username,
                "ip_address": s.ip_address or "",
                "user_agent": s.user_agent or "",
                "login_at": s.login_at.isoformat() if s.login_at else None,
                "last_seen_at": s.last_seen_at.isoformat() if s.last_seen_at else None,
                "logout_at": s.logout_at.isoformat() if s.logout_at else None,
                "ended_reason": s.ended_reason or "",
                "active": s.logout_at is None,
                "duration_seconds": duration_seconds,
            }
        )
    return jsonify(
        {
            "ok": True,
            "username": uname,
            "items": items,
            "total": total,
            "offset": offset,
            "limit": limit,
            "has_more": (offset + len(items)) < total,
            "sort_by": sort_by,
            "sort_order": sort_order,
        }
    )


@app.route('/api/admin/presence/export.csv', methods=["GET"])
def api_admin_presence_export_csv():
    _, admin_err = _require_admin()
    if admin_err:
        return admin_err
    username_filter = (request.args.get("username") or "").strip().lower()
    status_filter = (request.args.get("status") or "").strip().lower()
    rows, _ = _presence_rows_summary(username_filter=username_filter, status_filter=status_filter, limit=5000)

    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(
        [
            "username",
            "full_name",
            "email",
            "status",
            "last_seen_at",
            "active_sessions",
            "current_session_seconds",
            "last_session_seconds",
            "week_sessions_count",
            "week_total_seconds",
        ]
    )
    for r in rows:
        writer.writerow(
            [
                r.get("username", ""),
                r.get("full_name", ""),
                r.get("email", ""),
                r.get("status", ""),
                r.get("last_seen_at", ""),
                r.get("active_sessions", 0),
                r.get("current_session_seconds", 0),
                "" if r.get("last_session_seconds") is None else r.get("last_session_seconds"),
                r.get("week_sessions_count", 0),
                r.get("week_total_seconds", 0),
            ]
        )
    csv_body = out.getvalue()
    return Response(
        csv_body,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=presence_summary.csv"},
    )


@app.route('/api/changes', methods=["GET"])
def api_changes_list():
    _ensure_db_schema()
    actor, _ = _extract_actor_role()
    with db_session() as session:
        items = list_change_requests(session)
        ids = [int(i["id"]) for i in items if i.get("id") is not None]
        summary_map: Dict[int, Dict[str, int]] = {}
        thread_map: Dict[int, List[ChangeDiscussionThread]] = {}
        if ids:
            thread_rows = session.execute(
                select(ChangeDiscussionThread).where(ChangeDiscussionThread.change_request_id.in_(ids))
            ).scalars().all()
            for thread in thread_rows:
                change_request_id = int(thread.change_request_id)
                bucket = summary_map.setdefault(int(change_request_id), {"threads_total": 0, "threads_blocking": 0})
                bucket["threads_total"] += 1
                if bool(thread.requires_resolution) and thread.status != "resolved":
                    bucket["threads_blocking"] += 1
                thread_map.setdefault(change_request_id, []).append(thread)
        for item in items:
            sid = int(item.get("id"))
            summary = summary_map.get(sid, {"threads_total": 0, "threads_blocking": 0})
            item["threads_total"] = summary["threads_total"]
            item["threads_blocking"] = summary["threads_blocking"]
            actor_mentions = 0
            actor_needs_response = 0
            if actor:
                for thread in thread_map.get(sid, []):
                    if thread.status == "needs_author_response":
                        cr_author = (item.get("created_by") or "").strip()
                        if actor == cr_author:
                            actor_needs_response += 1
                    for msg in thread.messages or []:
                        if actor in _extract_mentions(msg.body or ""):
                            actor_mentions += 1
            item["my_mentions"] = actor_mentions
            item["my_needs_response"] = actor_needs_response
    return jsonify({"ok": True, "items": items})


@app.route('/api/changes', methods=["POST"])
def api_changes_create():
    _ensure_db_schema()
    data = request.get_json(silent=True) or {}
    actor, _ = _extract_actor_role(data)
    title = (data.get("title") or "Change request").strip()
    note = (data.get("note") or "").strip()
    submit_comment = (data.get("submit_comment") or "").strip()
    merge_mode = (data.get("merge_mode") or "append").strip()
    payload = data.get("payload") or {}
    target_domain = data.get("target_domain")
    target_skill = data.get("target_skill")
    with db_session() as session:
        current = load_unified_from_db(session, literature=load_literature_map())
        merged = merge_upload_into_source(
            current,
            payload,
            merge_mode=merge_mode,
            target_domain=target_domain,
            target_skill=target_skill,
        )
        revision_payload = build_revision_payload(
            base_snapshot=current,
            upload_payload=payload,
            proposed_snapshot=merged,
            merge_mode=merge_mode,
            target_domain=target_domain,
            target_skill=target_skill,
        )
        cr = create_change_request(
            session=session,
            title=title,
            merge_mode=merge_mode,
            payload=revision_payload,
            staging_batch_id=revision_payload.get("staging_batch_id"),
            created_by=actor,
            target_domain=target_domain,
            target_skill=target_skill,
            initial_note=note or "initial revision",
        )
        approval_set_status(session, cr.id, "submitted", actor=actor, comment=submit_comment or "Created and submitted")
        _notify_cr_submitted(session, cr)
        change_id = cr.id
    return jsonify({"ok": True, "id": change_id})


@app.route('/api/changes/<int:change_id>', methods=["GET"])
def api_changes_get(change_id):
    _ensure_db_schema()
    with db_session() as session:
        details = get_change_request_details(session, change_id)
        thread_rows = session.execute(
            select(ChangeDiscussionThread).where(ChangeDiscussionThread.change_request_id == change_id)
        ).scalars().all()
    if not details:
        return jsonify({"ok": False, "error": "Change request not found"}), 404
    details["discussion_summary"] = {
        "threads_total": len(thread_rows),
        "threads_blocking": len([t for t in thread_rows if t.requires_resolution and t.status != "resolved"]),
    }
    return jsonify({"ok": True, "change": details})


@app.route('/api/changes/<int:change_id>/discussion', methods=["GET"])
def api_changes_discussion(change_id: int):
    _ensure_db_schema()
    auth, auth_err = _require_authenticated()
    if auth_err:
        return auth_err
    with db_session() as session:
        cr = session.get(ChangeRequest, change_id)
        if not cr:
            return jsonify({"ok": False, "error": "Change request not found"}), 404
        rows = session.execute(
            select(ChangeDiscussionThread).where(ChangeDiscussionThread.change_request_id == change_id)
        ).scalars().all()
        status_filter = (request.args.get("status") or "").strip().lower()
        mentions_only = (request.args.get("mentions_only") or "").strip().lower() in ("1", "true", "yes")
        payload_threads = []
        for t in sorted(rows, key=lambda x: x.id or 0):
            if status_filter and status_filter != "all" and t.status != status_filter:
                continue
            serialized = _serialize_discussion_thread(t)
            mentioned = any(auth["actor"] in _extract_mentions(m.get("body") or "") for m in serialized.get("messages") or [])
            needs_author_response = (t.status == "needs_author_response" and (cr.created_by or "") == auth["actor"])
            serialized["mentioned_me"] = mentioned
            serialized["needs_my_response"] = needs_author_response
            if mentions_only and not mentioned and not needs_author_response:
                continue
            payload_threads.append(serialized)
    return jsonify({"ok": True, "requested_by": auth["actor"], "threads": payload_threads})


@app.route('/api/changes/<int:change_id>/timeline', methods=["GET"])
def api_changes_timeline(change_id: int):
    _ensure_db_schema()
    auth, auth_err = _require_authenticated()
    if auth_err:
        return auth_err
    with db_session() as session:
        cr = session.get(ChangeRequest, change_id)
        if not cr:
            return jsonify({"ok": False, "error": "Change request not found"}), 404
        revisions = session.execute(
            select(ChangeRevision).where(ChangeRevision.change_request_id == change_id).order_by(ChangeRevision.revision_no.asc())
        ).scalars().all()
        decisions = session.execute(
            select(ApprovalDecision).where(ApprovalDecision.change_request_id == change_id).order_by(ApprovalDecision.id.asc())
        ).scalars().all()
        threads = session.execute(
            select(ChangeDiscussionThread).where(ChangeDiscussionThread.change_request_id == change_id)
        ).scalars().all()
        events: List[Dict[str, Any]] = []
        events.append({
            "kind": "change_created",
            "at": cr.created_at.isoformat() if cr.created_at else None,
            "actor": cr.created_by,
            "body": f"Change request created: {cr.title}",
        })
        for rev in revisions:
            events.append({
                "kind": "revision",
                "at": rev.created_at.isoformat() if rev.created_at else None,
                "actor": rev.created_by,
                "body": f"Revision #{rev.revision_no}. {rev.note or ''}".strip(),
            })
        for d in decisions:
            events.append({
                "kind": "status",
                "at": d.created_at.isoformat() if d.created_at else None,
                "actor": d.actor,
                "body": f"Status -> {d.decision}. {d.comment or ''}".strip(),
            })
        for t in threads:
            events.append({
                "kind": "thread",
                "at": t.created_at.isoformat() if t.created_at else None,
                "actor": t.created_by,
                "body": f"Thread #{t.id}: {t.subject}",
            })
            for m in sorted(t.messages or [], key=lambda x: x.id or 0):
                events.append({
                    "kind": "thread_message",
                    "at": m.created_at.isoformat() if m.created_at else None,
                    "actor": m.author,
                    "body": f"[thread #{t.id}] {m.body}",
                    "mentions_me": auth["actor"] in _extract_mentions(m.body or ""),
                })
        events.sort(key=lambda e: e.get("at") or "")
    return jsonify({"ok": True, "requested_by": auth["actor"], "events": events})


@app.route('/api/changes/<int:change_id>/discussion/threads', methods=["POST"])
def api_changes_discussion_create_thread(change_id: int):
    _ensure_db_schema()
    auth, auth_err = _require_authenticated()
    if auth_err:
        return auth_err
    data = request.get_json(silent=True) or {}
    subject = (data.get("subject") or "").strip() or "Комментарий к изменению"
    body = (data.get("body") or "").strip()
    requires_resolution = bool(data.get("requires_resolution", True))
    if not body:
        return jsonify({"ok": False, "error": "body required"}), 400
    with db_session() as session:
        cr = session.get(ChangeRequest, change_id)
        if not cr:
            return jsonify({"ok": False, "error": "Change request not found"}), 404
        thread = ChangeDiscussionThread(
            change_request_id=change_id,
            subject=subject,
            status="open",
            requires_resolution=requires_resolution,
            created_by=auth["actor"],
            created_role=auth["role"],
            resolved_by=None,
            resolved_at=None,
        )
        session.add(thread)
        session.flush()
        session.add(
            ChangeDiscussionMessage(
                thread_id=thread.id,
                author=auth["actor"],
                author_role=auth["role"],
                body=body,
                kind="comment",
            )
        )
        _notify_mentions(session, change_id=change_id, actor=auth["actor"], text_value=body)
        cr.updated_at = thread.updated_at
        thread_id = thread.id
    return jsonify({"ok": True, "thread_id": thread_id})


@app.route('/api/changes/<int:change_id>/discussion/threads/<int:thread_id>/messages', methods=["POST"])
def api_changes_discussion_add_message(change_id: int, thread_id: int):
    _ensure_db_schema()
    auth, auth_err = _require_authenticated()
    if auth_err:
        return auth_err
    data = request.get_json(silent=True) or {}
    body = (data.get("body") or "").strip()
    if not body:
        return jsonify({"ok": False, "error": "body required"}), 400
    with db_session() as session:
        thread = session.get(ChangeDiscussionThread, thread_id)
        if not thread or thread.change_request_id != change_id:
            return jsonify({"ok": False, "error": "Thread not found"}), 404
        session.add(
            ChangeDiscussionMessage(
                thread_id=thread.id,
                author=auth["actor"],
                author_role=auth["role"],
                body=body,
                kind="comment",
            )
        )
        _notify_mentions(session, change_id=change_id, actor=auth["actor"], text_value=body)
        # Any new reply reopens resolution loop unless it was already fully resolved by admin decision.
        if thread.status != "resolved":
            thread.status = "open"
            thread.resolved_by = None
            thread.resolved_at = None
    return jsonify({"ok": True})


@app.route('/api/changes/<int:change_id>/discussion/threads/<int:thread_id>/status', methods=["POST"])
def api_changes_discussion_set_status(change_id: int, thread_id: int):
    _ensure_db_schema()
    data = request.get_json(silent=True) or {}
    actor, admin_err = _require_admin(data)
    if admin_err:
        return admin_err
    action = (data.get("action") or "").strip().lower()
    comment = (data.get("comment") or "").strip()
    mapping = {
        "accept": "resolved",
        "resolve": "resolved",
        "reject": "needs_author_response",
        "request_info": "needs_author_response",
        "reopen": "open",
    }
    new_status = mapping.get(action)
    if new_status not in DISCUSSION_THREAD_STATUSES:
        return jsonify({"ok": False, "error": "Invalid action"}), 400
    if action in {"reject", "request_info"} and not comment:
        return jsonify({"ok": False, "error": "comment required for reject/request_info"}), 400
    with db_session() as session:
        thread = session.get(ChangeDiscussionThread, thread_id)
        if not thread or thread.change_request_id != change_id:
            return jsonify({"ok": False, "error": "Thread not found"}), 404
        thread.status = new_status
        if new_status == "resolved":
            thread.resolved_by = actor
            thread.resolved_at = datetime.now(timezone.utc)
        else:
            thread.resolved_by = None
            thread.resolved_at = None
        session.add(
            ChangeDiscussionMessage(
                thread_id=thread.id,
                author=actor,
                author_role="admin",
                body=comment or f"Status changed: {new_status}",
                kind="status",
            )
        )
        _notify_mentions(session, change_id=change_id, actor=actor, text_value=comment or "")
        if new_status == "needs_author_response":
            cr = session.get(ChangeRequest, change_id)
            if cr:
                _notify_cr_status_to_author(
                    session,
                    cr,
                    status="needs_author_response",
                    actor=actor,
                    comment=comment or "Admin requested additional details in discussion",
                )
    return jsonify({"ok": True, "status": new_status})


@app.route('/api/changes/<int:change_id>/revise', methods=["POST"])
def api_changes_revise(change_id):
    _ensure_db_schema()
    data = request.get_json(silent=True) or {}
    payload = data.get("payload") or {}
    actor, _ = _extract_actor_role(data)
    note = (data.get("note") or "").strip()
    with db_session() as session:
        cr = session.get(ChangeRequest, change_id)
        if not cr:
            return jsonify({"ok": False, "error": "Change request not found"}), 404
        current = load_unified_from_db(session, literature=load_literature_map())
        merged = merge_upload_into_source(
            current,
            payload,
            merge_mode=cr.merge_mode,
            target_domain=cr.target_domain,
            target_skill=cr.target_skill,
        )
        revision_payload = build_revision_payload(
            base_snapshot=current,
            upload_payload=payload,
            proposed_snapshot=merged,
            merge_mode=cr.merge_mode,
            target_domain=cr.target_domain,
            target_skill=cr.target_skill,
        )
        rev = add_revision(
            session,
            change_id,
            payload=revision_payload,
            actor=actor,
            note=note,
            staging_batch_id=revision_payload.get("staging_batch_id"),
        )
    return jsonify({"ok": True, "revision_no": rev.revision_no})


@app.route('/api/changes/<int:change_id>/status', methods=["POST"])
def api_changes_status(change_id):
    _ensure_db_schema()
    data = request.get_json(silent=True) or {}
    actor, admin_err = _require_admin(data)
    if admin_err:
        return admin_err
    status = (data.get("status") or "").strip()
    comment = (data.get("comment") or "").strip()
    with db_session() as session:
        cr = approval_set_status(session, change_id, status=status, actor=actor, comment=comment)
        if not cr:
            return jsonify({"ok": False, "error": "Invalid status or change request not found"}), 400
        if status in {"approved", "rejected", "in_review"}:
            _notify_cr_status_to_author(session, cr, status=status, actor=actor, comment=comment)
    return jsonify({"ok": True, "status": status})


@app.route('/api/changes/<int:change_id>/apply', methods=["POST"])
def api_changes_apply(change_id):
    _ensure_db_schema()
    data = request.get_json(silent=True) or {}
    actor, admin_err = _require_admin(data)
    if admin_err:
        return admin_err
    with db_session() as session:
        cr = session.get(ChangeRequest, change_id)
        if not cr:
            return jsonify({"ok": False, "error": "Change request not found"}), 404
        blocking_threads = session.execute(
            select(func.count(ChangeDiscussionThread.id)).where(
                ChangeDiscussionThread.change_request_id == change_id,
                ChangeDiscussionThread.requires_resolution == True,
                ChangeDiscussionThread.status != "resolved",
            )
        ).scalar_one()
        if int(blocking_threads or 0) > 0:
            return jsonify({"ok": False, "error": "There are unresolved discussion threads"}), 409
        if cr.status != "approved":
            return jsonify({"ok": False, "error": "Only approved changes can be applied"}), 409
        payload = get_latest_payload(session, change_id) or {}
        proposed = payload.get("proposed_snapshot") if isinstance(payload, dict) else None
        if not isinstance(proposed, dict):
            upload_payload = payload.get("upload_payload") if isinstance(payload, dict) else payload
            current = load_unified_from_db(session, literature=load_literature_map())
            proposed = merge_upload_into_source(
                current,
                upload_payload or {},
                merge_mode=cr.merge_mode,
                target_domain=cr.target_domain,
                target_skill=cr.target_skill,
            )
        prop = proposed if isinstance(proposed, dict) else {}
        # Полный снимок из CR: дерево + META_KEYS (ui_config с matrix_levels / matrix_column_schema и т.д.).
        literature = load_literature_map()
        current_unified = load_unified_from_db(session, literature=literature)
        current_unified["domains"] = prop.get("domains") or []
        current_unified["nodes"] = deepcopy(prop.get("nodes") or [])
        for k in META_KEYS:
            if k in prop and prop[k] is not None:
                current_unified[k] = deepcopy(prop[k])
        replace_unified_in_db(session, current_unified)
        cr.applied = True
        approval_set_status(session, change_id, "applied", actor=actor, comment="Applied to storage")
        _notify_cr_status_to_author(session, cr, status="applied", actor=actor, comment="Applied to storage")
    _invalidate_caches()
    _ensure_data_loaded()
    return jsonify({"ok": True, "applied": True})


# ----- ЛИТЕРАТУРА: каталог, привязка к листам -----

@app.route('/literature')
def literature_page():
    return render_template('literature.html')

@app.route('/api/literature')
def api_literature_list():
    """Список литературы с привязкой к компетенциям (листам)."""
    meta = get_meta()
    lit = meta.get('literature', {})
    templates = meta.get('action_templates', {})
    template_to_lit = {}
    for tid, t in templates.items():
        for rid in t.get('resource_ids', []):
            template_to_lit.setdefault(rid, []).append({"template_id": tid, "name": t.get('name', tid)})
    leaf_by_lit = _literature_linked_leaves(meta)
    out = []
    for rid, item in lit.items():
        local_path = item.get("local_path") or item.get("file_path", "")
        if local_path:
            abs_local = local_path if os.path.isabs(local_path) else os.path.join(BASE_DIR, local_path)
            if not os.path.isfile(abs_local):
                local_path = ""
        out.append({
            "id": rid,
            "title": item.get("title", rid),
            "chapter": item.get("chapter", ""),
            "pages": item.get("pages", ""),
            "url": item.get("url", ""),
            "description": item.get("description", ""),
            "local_path": local_path,
            "linked_templates": template_to_lit.get(rid, []),
            "linked_leaves": leaf_by_lit.get(str(rid), []),
        })
    out.extend(_matrix_only_source_catalog(meta))
    return jsonify(out)

@app.route('/api/literature', methods=['POST'])
def api_literature_add():
    """Ручное добавление литературы."""
    global _meta
    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    if not title:
        return jsonify({"error": "title required"}), 400
    meta = get_meta()
    lit = meta.setdefault('literature', {})
    lid = slugify(title)[:40] + '_' + hashlib.md5(title.encode()).hexdigest()[:6]
    if lid in lit:
        return jsonify({"error": "already exists", "id": lid}), 409
    _create_version_backup("literature_add", f"Добавление источника: {title}")
    lit[lid] = {
        "title": title,
        "chapter": (data.get('chapter') or '').strip(),
        "pages": (data.get('pages') or '').strip(),
        "url": (data.get('url') or '').strip(),
        "description": (data.get('description') or '').strip(),
    }
    save_meta(meta)
    return jsonify({"id": lid, "title": title})


@app.route('/api/literature/upload', methods=['POST'])
def api_literature_upload():
    """Загрузка физического файла (PDF и т.д.) в data/library с созданием записи литературы."""
    global _meta
    if "file" not in request.files:
        return jsonify({"error": "файл не выбран"}), 400
    f = request.files["file"]
    if f.filename == "" or not f.filename:
        return jsonify({"error": "файл не выбран"}), 400
    title = (request.form.get("title") or "").strip() or PathLib(f.filename).stem
    chapter = (request.form.get("chapter") or "").strip()
    pages = (request.form.get("pages") or "").strip()
    description = (request.form.get("description") or "").strip()
    lib_dir = _literature_dir()
    orig = PathLib(f.filename)
    suffix = (orig.suffix.lower() or ".pdf")
    if not suffix.startswith("."):
        suffix = "." + suffix
    stem = re.sub(r"[^\w\-]", "_", orig.stem)[:60]
    filepath = os.path.join(lib_dir, stem + suffix)
    n = 0
    while os.path.exists(filepath):
        n += 1
        filepath = os.path.join(lib_dir, f"{stem}_{n}{suffix}")
    try:
        f.save(filepath)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    rel_path = os.path.relpath(filepath, BASE_DIR)
    meta = get_meta()
    lit = meta.setdefault("literature", {})
    lid = slugify(title)[:40] + "_" + hashlib.md5((title + rel_path).encode()).hexdigest()[:6]
    if lid in lit:
        lid = lid + "_" + hashlib.md5(rel_path.encode()).hexdigest()[:4]
    _create_version_backup("literature_upload", f"Загрузка файла литературы: {title}")
    lit[lid] = {
        "title": title,
        "chapter": chapter,
        "pages": pages,
        "url": "",
        "description": description,
        "local_path": rel_path,
    }
    if _is_db_mode():
        mongo_add_literature_file(lid, os.path.basename(filepath), rel_path, suffix.lstrip("."))
    save_meta(meta)
    return jsonify({"id": lid, "title": title, "local_path": rel_path})


@app.route('/api/literature/<lit_id>', methods=['PATCH'])
def api_literature_edit(lit_id):
    """Редактирование литературы: title/chapter/pages/url/description/local_path."""
    global _meta
    data = request.get_json() or {}
    meta = get_meta()
    lit = meta.get('literature', {})
    if lit_id not in lit:
        return jsonify({"error": "literature not found"}), 404
    _create_version_backup("literature_edit", f"Редактирование литературы: {lit_id}")
    item = lit[lit_id]
    if 'title' in data:
        item['title'] = (data.get('title') or '').strip()
    if 'chapter' in data:
        item['chapter'] = (data.get('chapter') or '').strip()
    if 'pages' in data:
        item['pages'] = (data.get('pages') or '').strip()
    if 'url' in data:
        item['url'] = (data.get('url') or '').strip()
    if 'description' in data:
        item['description'] = (data.get('description') or '').strip()
    if 'local_path' in data:
        item['local_path'] = (data.get('local_path') or '').strip()
    save_meta(meta)
    return jsonify({"id": lit_id, "title": item.get("title", lit_id)})


@app.route('/api/literature/<lit_id>', methods=['DELETE'])
def api_literature_delete(lit_id):
    """Удаление литературы: удаляет из каталога и убирает привязки (resource_ids) из всех шаблонов."""
    global _meta
    meta = get_meta()
    lit = meta.get('literature', {})
    if lit_id not in lit:
        return jsonify({"error": "literature not found"}), 404
    _create_version_backup("literature_delete", f"Удаление литературы: {lit_id}")
    templates = meta.get('action_templates', {})
    removed = 0
    for tid, t in templates.items():
        rids = t.get('resource_ids', [])
        if lit_id in rids:
            t['resource_ids'] = [r for r in rids if r != lit_id]
            removed += 1
    del lit[lit_id]
    if _is_db_mode():
        mongo_delete_literature_item(lit_id)
    save_meta(meta)
    return jsonify({"id": lit_id, "removed_from_templates": removed})


@app.route('/api/literature/<lit_id>/link', methods=['POST'])
def api_literature_link(lit_id):
    """Привязка литературы к листам (по path). Добавляет lit_id в resource_ids шаблонов этих листов."""
    global _meta
    data = request.get_json() or {}
    leaf_paths = data.get('leaf_paths') or []
    if not leaf_paths:
        return jsonify({"error": "leaf_paths required"}), 400
    meta = get_meta()
    if lit_id not in meta.get('literature', {}):
        return jsonify({"error": "literature not found"}), 404
    _create_version_backup("literature_link", f"Привязка литературы: {lit_id}")
    templates = meta.setdefault('action_templates', {})
    tree = get_tree()
    updated = 0
    for path_str in leaf_paths:
        try:
            path = [int(x) for x in str(path_str).strip("/").split("/") if x.strip()]
        except ValueError:
            continue
        node = get_node_by_path(tree, path)
        if not node or node.get('children'):
            continue
        tid = node.get('template_id')
        if not tid or tid not in templates:
            continue
        rids = templates[tid].setdefault('resource_ids', [])
        if lit_id not in rids:
            rids.append(lit_id)
            updated += 1
    save_meta(meta)
    return jsonify({"updated": updated})


@app.route('/api/literature/<lit_id>/unlink', methods=['POST'])
def api_literature_unlink(lit_id):
    """Отвязка литературы от листов (по path): удаляет lit_id из resource_ids шаблонов."""
    global _meta
    data = request.get_json() or {}
    leaf_paths = data.get('leaf_paths') or []
    if not leaf_paths:
        return jsonify({"error": "leaf_paths required"}), 400
    meta = get_meta()
    if lit_id not in meta.get('literature', {}):
        return jsonify({"error": "literature not found"}), 404
    _create_version_backup("literature_unlink", f"Отвязка литературы: {lit_id}")
    templates = meta.setdefault('action_templates', {})
    tree = get_tree()
    updated = 0
    for path_str in leaf_paths:
        try:
            path = [int(x) for x in str(path_str).strip("/").split("/") if x.strip()]
        except ValueError:
            continue
        node = get_node_by_path(tree, path)
        if not node or node.get('children'):
            continue
        tid = node.get('template_id')
        if not tid or tid not in templates:
            continue
        rids = templates[tid].setdefault('resource_ids', [])
        if lit_id in rids:
            templates[tid]['resource_ids'] = [rid for rid in rids if rid != lit_id]
            updated += 1
    save_meta(meta)
    return jsonify({"updated": updated})


# Типы контента, которые считаем файлами для скачивания (не веб-страницы)
_DOWNLOADABLE_CONTENT_TYPES = (
    'application/pdf', 'application/octet-stream', 'application/x-pdf',
    'application/msword', 'application/vnd.openxmlformats-officedocument.',
    'application/zip', 'application/x-rar', 'application/epub+zip',
    'image/', 'audio/', 'video/', 'text/csv', 'text/plain',
)


def _is_downloadable_content_type(ct):
    """Проверяет, является ли Content-Type скачиваемым файлом (не веб-страница)."""
    if not ct:
        return False
    ct_lower = ct.lower().split(';')[0].strip()
    if ct_lower.startswith(('text/html', 'application/xhtml', 'text/xml')):
        return False
    for prefix in _DOWNLOADABLE_CONTENT_TYPES:
        if ct_lower.startswith(prefix):
            return True
    return False


@app.route('/api/literature/<lit_id>/download', methods=['POST'])
def api_literature_download(lit_id):
    """Скачивает файл по URL в каталог data/library и проставляет local_path у записи литературы."""
    import urllib.request
    import urllib.parse
    data = request.get_json() or {}
    url = (data.get('url') or '').strip()
    if not url:
        return jsonify({"error": "url required"}), 400
    meta = get_meta()
    if lit_id not in meta.get('literature', {}):
        return jsonify({"error": "literature not found"}), 404
    _create_version_backup("literature_download", f"Скачивание литературы по URL: {lit_id}")
    lib_dir = _literature_dir()
    try:
        import ssl
        cfg = load_app_config()
        ssl_verify = cfg.get("ssl_verify", True)
        if os.environ.get("DE_MATRIX_SSL_VERIFY", "").lower() in ("0", "false", "no"):
            ssl_verify = False
        ssl_ca_bundle = cfg.get("ssl_ca_bundle") or os.environ.get("DE_MATRIX_SSL_CA_BUNDLE")

        if not ssl_verify:
            ctx = ssl._create_unverified_context()
        elif ssl_ca_bundle and os.path.isfile(ssl_ca_bundle):
            ctx = ssl.create_default_context(cafile=ssl_ca_bundle)
        else:
            # Системные сертификаты (включая корпоративные CA в корпоративных сетях)
            ctx = ssl.create_default_context()

        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
            content_type = resp.headers.get('Content-Type', '')
            if not _is_downloadable_content_type(content_type):
                resp.read()  # consume and close
                return jsonify({
                    "error": "Файлов для скачивания не найдено. Ссылка ведёт на веб-страницу.",
                    "preview_url": url,
                }), 400
            content = resp.read()
        ext = os.path.splitext(urllib.parse.urlparse(url).path)[1] or ".pdf"
        safe_id = re.sub(r'[^\w\-]', '_', lit_id)[:50]
        filename = f"{safe_id}{ext}"
        filepath = os.path.join(lib_dir, filename)
        with open(filepath, 'wb') as f:
            f.write(content)
        rel_path = os.path.relpath(filepath, BASE_DIR)
        meta['literature'][lit_id]['local_path'] = rel_path
        save_meta(meta)
        return jsonify({"local_path": rel_path})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/domains')
def api_domains():
    """Структура корней и детей первого уровня для догрузки."""
    _ensure_data_loaded()
    matrix = get_matrix() or {}
    nodes = matrix.get("nodes") or []
    out = []
    for n in nodes:
        if not isinstance(n, dict):
            continue
        ch = n.get("children") or []
        out.append(
            {
                "name": n.get("name", ""),
                "skills": [c.get("name", "") for c in ch if isinstance(c, dict)],
            }
        )
    return jsonify({"ok": True, "domains": out})


@app.route('/api/domain/<int:domain_idx>')
def api_domain(domain_idx):
    """Полные данные домена для вью дерева (слева направо)."""
    matrix = get_matrix()
    nodes = (matrix or {}).get("nodes") or []
    if domain_idx < 0 or domain_idx >= _matrix_root_count(matrix):
        return jsonify({"ok": False, "error": "Domain not found"}), 404
    root = nodes[domain_idx]
    if not isinstance(root, dict):
        return jsonify({"ok": False, "error": "Domain not found"}), 404
    domain_color = get_domain_color(root.get("name", ""))
    domain_icon = get_domain_icon(root.get("name", ""), domain_idx)
    out = {
        "ok": True,
        "domain": {
            "index": domain_idx,
            "name": root.get("name", ""),
            "color": domain_color,
            "icon": domain_icon,
            "skills": [],
        },
        }
    for si, s in enumerate(root.get("children") or []):
        if not isinstance(s, dict):
            continue
        skill_color = get_skill_color(s.get("name", ""), domain_color, si)
        skill_icon = get_skill_icon(s.get("name", ""), si)
        out["domain"]["skills"].append(
            {
            "index": si,
            "name": s.get("name", ""),
            "description": s.get("description", ""),
            "responsible": s.get("responsible", ""),
            "level_sticker": s.get("level_sticker", ""),
            "color": skill_color,
            "icon": skill_icon,
                "actions": [],
            }
        )
    return jsonify(out)


@app.route('/api/sources')
def api_sources():
    """DB-only источник данных для рантайма."""
    return jsonify({
        "source_dir": "db://postgres",
        "files": ["db://postgres"],
        "current": "db://postgres",
    })


@app.route('/api/import/template')
def api_import_template():
    """Скачивание пустого шаблона Excel для импорта (единый формат с экспортом)."""
    _, admin_err = _require_admin()
    if admin_err:
        return admin_err
    try:
        from openpyxl import Workbook
        from io import BytesIO
    except ImportError:
        return jsonify({"error": "openpyxl не установлен (pip install openpyxl)"}), 500
    wb = Workbook()
    ws = wb.active
    ws.title = "Матрица навыков"
    ws.append(["Domain", "Skill", "Action", "Subaction", "Description", "Template ID", "Level Tag", "Review Questions"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="matrix_import_template.xlsx",
    )


@app.route('/api/import/template/unified')
def api_import_template_unified():
    """Шаблон unified relational: первая строка 1:1 с импортом (header из matrix_column_schema) или синтетика."""
    _, admin_err = _require_admin()
    if admin_err:
        return admin_err
    try:
        from openpyxl import Workbook
        from io import BytesIO
    except ImportError:
        return jsonify({"error": "openpyxl не установлен (pip install openpyxl)"}), 500
    meta = get_meta()
    ui = meta.get("ui_config") or {}
    schema = effective_matrix_column_schema(ui)
    header = [matrix_roundtrip_header_cell(e, ui) for e in schema]
    wb = Workbook()
    ws = wb.active
    ws.title = "Unified_Relational_Span"
    ws.append(header)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="matrix_unified_template.xlsx",
    )


@app.route('/api/source/upload/preview', methods=["POST"])
def api_source_upload_preview():
    """Предпросмотр догрузки: парсинг файла без сохранения, возврат preview + validation."""
    _, admin_err = _require_admin()
    if admin_err:
        return admin_err
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Файл не передан"}), 400
    f = request.files["file"]
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "Файл не выбран"}), 400
    ext = (os.path.splitext(f.filename)[1] or "").lower()
    if ext not in (".json", ".xlsx", ".xls"):
        return jsonify({"ok": False, "error": "Поддерживаются только JSON и Excel"}), 400

    try:
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name
        try:
            if ext == ".json":
                with open(tmp_path, "r", encoding="utf-8") as fp:
                    upload_data = json.load(fp)
            else:
                upload_data = load_excel_for_matrix_import(tmp_path)
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "validation": {"ok": False, "errors": [str(e)]}}), 400

    from core.loaders import _normalize_unified
    from core.schema import validate_source
    upload_data = _normalize_unified(upload_data)
    vr = validate_source(upload_data)

    def _preview_level_field(node: Dict[str, Any]) -> str:
        tags = node.get("level_tags") if isinstance(node.get("level_tags"), list) else []
        if tags:
            return ", ".join(str(x) for x in tags if str(x).strip())
        return str(node.get("level_tag") or "").strip()

    def _preview_leaf_keys(node: Dict[str, Any]) -> str:
        lv = node.get("leaf_view")
        if not isinstance(lv, dict) or not lv:
            return ""
        keys = []
        for k, v in lv.items():
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            if isinstance(v, (list, dict)) and not v:
                continue
            keys.append(str(k))
        return ", ".join(sorted(keys))

    # Preview: плоская таблица по листьям generic-дерева nodes
    preview_rows = []
    u_nodes_preview = upload_data.get("nodes") or []
    if u_nodes_preview:
        tprev = deepcopy(u_nodes_preview)
        assign_paths_to_generic_nodes(tprev)
        for leaf in collect_leaves(tprev):
            p = leaf.get("path") or []
            if len(p) < 3:
                continue
            anc = get_ancestors(tprev, p)
            chain = list(anc) + [leaf]
            names = [(x.get("name") or "").strip() for x in chain]
            d_name = names[0] if names else ""
            s_name = names[1] if len(names) > 1 else ""
            depth = len(names)
            if depth >= 4:
                action_text = names[-2]
                sub_text = names[-1]
            else:
                action_text = names[-1]
                sub_text = ""
            sk_node = chain[1] if len(chain) > 1 else {}
            skill_resp = str(sk_node.get("responsible") or "").strip()
            skill_sticker = str(sk_node.get("level_sticker") or "").strip()
            preview_rows.append(
                {
                    "domain": d_name,
                    "skill": s_name,
                    "skill_responsible": skill_resp,
                    "skill_level_sticker": skill_sticker,
                    "action": action_text,
                    "subaction": sub_text,
                    "template_id": leaf.get("template_id"),
                    "level_tag": leaf.get("level_tag") or "",
                    "level_tags": _preview_level_field(leaf),
                    "leaf_view_keys": _preview_leaf_keys(leaf),
                    "review_questions": "; ".join(str(q) for q in (leaf.get("review_questions") or []) if q),
                }
            )

    ui_cfg = upload_data.get("ui_config") if isinstance(upload_data.get("ui_config"), dict) else {}
    mcs = ui_cfg.get("matrix_column_schema")
    preview_unified: Optional[Dict[str, Any]] = None
    if isinstance(mcs, list) and len(mcs) > 0:
        try:
            u_nodes = upload_data.get("nodes") or []
            u_domains = upload_data.get("domains") or []
            u_headers, u_rows = build_unified_export_table(
                u_domains,
                ui_cfg,
                nodes=(u_nodes or None),
                include_header_tags=False,
            )
            if u_headers and u_rows:
                max_rows = 250
                preview_unified = {
                    "headers": u_headers,
                    "rows": u_rows[:max_rows],
                    "truncated": len(u_rows) > max_rows,
                    "total_rows": len(u_rows),
                }
        except Exception:
            preview_unified = None

    preview_context = {
        "has_ui_config": bool(ui_cfg),
        "matrix_levels_count": len(ui_cfg.get("matrix_levels") or []),
        "matrix_column_schema_count": len(ui_cfg.get("matrix_column_schema") or []),
        "unified_preview": bool(preview_unified),
        "note": (
            "Для unified с matrix_column_schema порядок колонок и подписи превью соответствуют шапке файла "
            "(текст до скобок с тегами). Полная строка шапки сохраняется в схеме и уходит в шаблон/экспорт XLSX. "
            "Иначе — укороченный плоский вид (домен, навык, …)."
        ),
    }

    return jsonify({
        "ok": True,
        "preview": preview_rows,
        "preview_unified": preview_unified,
        "preview_context": preview_context,
        "validation": vr.to_dict(),
        "matrix_roots": len(upload_data.get("nodes") or []),
        "domains_count": len(upload_data.get("nodes") or []),
    })


@app.route('/api/source/upload', methods=["POST"])
def api_source_upload():
    """Догрузка данных из JSON/Excel только в approval pipeline (submit-only)."""
    _, admin_err = _require_admin()
    if admin_err:
        return admin_err
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Файл не передан (ожидается поле 'file')"}), 400
    f = request.files["file"]
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "Файл не выбран"}), 400
    ext = (os.path.splitext(f.filename)[1] or "").lower()
    if ext not in (".json", ".xlsx", ".xls"):
        return jsonify({"ok": False, "error": "Поддерживаются только JSON и Excel (.json, .xlsx, .xls)"}), 400

    merge_mode = (request.form.get("merge_mode") or "append").strip()
    # Для Excel-импорта в admin-вкладке используем сценарий переинициализации матрицы.
    if ext in (".xlsx", ".xls"):
        merge_mode = "replace_all"
    target_domain = (request.form.get("target_domain") or "").strip() or None
    target_skill = (request.form.get("target_skill") or "").strip() or None
    if merge_mode not in ("append", "append_to_domain", "append_to_skill", "replace_domain", "replace_skill", "replace_all"):
        merge_mode = "append"
    if merge_mode == "append_to_domain" and not target_domain:
        return jsonify({"ok": False, "error": "Для режима «В домен» укажите target_domain"}), 400
    if merge_mode == "append_to_skill" and (not target_domain or not target_skill):
        return jsonify({"ok": False, "error": "Для режима «В навык» укажите target_domain и target_skill"}), 400

    try:
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name
        try:
            if ext == ".json":
                with open(tmp_path, "r", encoding="utf-8") as fp:
                    upload_data = json.load(fp)
            else:
                upload_data = load_excel_for_matrix_import(tmp_path)
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    except Exception as e:
        return jsonify({"ok": False, "error": f"Ошибка чтения файла: {e}"}), 400

    try:
        _ensure_db_schema()
        actor, _ = _extract_actor_role()
        with db_session() as session:
            current_unified = load_unified_from_db(session, literature=load_literature_map())
            current_nodes = load_matrix_nodes_nested(session)
            staging_batch = create_staging_batch(
                session,
                source_filename=f.filename or "upload",
                merge_mode=merge_mode,
                created_by=actor,
                payload=upload_data,
                target_domain=target_domain,
                target_skill=target_skill,
            )
            staging_batch_id = staging_batch.id
            staging_projection = load_staging_tree_projection(session, staging_batch_id)
            merged = merge_upload_into_source(
                current_unified, upload_data,
                merge_mode=merge_mode,
                target_domain=target_domain,
                target_skill=target_skill,
            )
        base_snap = {"domains": [], "nodes": current_nodes}
        revision_payload = build_revision_payload(
            base_snapshot=base_snap,
            upload_payload=upload_data,
            proposed_snapshot=merged,
            merge_mode=merge_mode,
            target_domain=target_domain,
            target_skill=target_skill,
        )
        revision_payload["staging_batch_id"] = staging_batch_id
        revision_payload["staging_tree"] = staging_projection
        title = f"Upload {f.filename}"
        with db_session() as session:
            cr = create_change_request(
                session=session,
                title=title,
                merge_mode=merge_mode,
                payload=revision_payload,
                staging_batch_id=staging_batch_id,
                created_by=actor,
                target_domain=target_domain,
                target_skill=target_skill,
            )
            approval_set_status(session, cr.id, "submitted", actor=actor, comment="Uploaded and submitted for review")
            _notify_cr_submitted(session, cr)
            change_id = cr.id
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({
        "ok": True,
        "queued_for_approval": True,
        "message": "Изменения отправлены на апрув и еще не применены к матрице",
        "change_id": change_id,
        "staging_batch_id": staging_batch_id,
        "merge_mode": merge_mode,
        "approval_required": storage_approval_required(),
    })


@app.route('/api/source/load', methods=["POST"])
def api_source_load():
    """DB-only рантайм: выбор file source отключен."""
    return jsonify({"ok": False, "error": "DB-only mode: source file switching is disabled"}), 400


@app.route('/api/backups')
def api_backups():
    """Список бэкапов конфига и источника."""
    if _is_db_mode():
        return jsonify({
            "ok": True,
            "backups": [],
            "mode": "db",
            "message": "Use scripts/db_backup.sh for PostgreSQL+Mongo backups",
        })
    source_file = _current_source_for_backup()
    if source_file:
        ensure_stable_backup(
            base_dir=PathLib(BASE_DIR),
            config_dir=PathLib(BASE_DIR) / "config",
            source_dir=PathLib(_source_dir_path()),
            source_filename=source_file,
            checkpoint_path=_checkpoint_path(),
        )
    backups = list_backups(PathLib(BASE_DIR))
    return jsonify({"ok": True, "backups": backups})


@app.route('/api/backups/<backup_id>/compatibility')
def api_backup_compatibility(backup_id):
    """Проверка совместимости бэкапа с текущей схемой."""
    if _is_db_mode():
        return jsonify({
            "ok": False,
            "error": "DB mode: file backup compatibility is not applicable",
        }), 400
    compatible, warning = check_backup_compatibility(PathLib(BASE_DIR), backup_id)
    return jsonify({
        "ok": True,
        "backup_id": backup_id,
        "compatible": compatible,
        "warning": warning,
    })


@app.route('/api/backups/<backup_id>/mark-stable', methods=["POST"])
def api_mark_stable(backup_id):
    """Помечает выбранный бэкап как стабильное состояние."""
    if _is_db_mode():
        return jsonify({
            "ok": False,
            "error": "DB mode: use database-level backup policies",
        }), 400
    backups = list_backups(PathLib(BASE_DIR))
    exists = any(b.get("id") == backup_id for b in backups)
    if not exists:
        return jsonify({"ok": False, "error": "Backup not found"}), 404
    if not set_stable_backup_id(PathLib(BASE_DIR), backup_id):
        return jsonify({"ok": False, "error": "Не удалось сохранить stable-состояние"}), 500
    return jsonify({"ok": True, "stable_backup_id": backup_id})


@app.route('/api/restore', methods=["POST"])
def api_restore():
    """Восстановление из бэкапа: перезапись config и source, перезагрузка данных."""
    if _is_db_mode():
        return jsonify({
            "ok": False,
            "error": "DB mode: use scripts/db_restore.sh",
        }), 400
    data = request.get_json() or {}
    backup_id = (data.get("backup_id") or data.get("id") or "").strip()
    force = bool(data.get("force"))
    if not backup_id:
        return jsonify({"ok": False, "error": "backup_id required"}), 400

    compatible, warning = check_backup_compatibility(PathLib(BASE_DIR), backup_id)
    if not compatible and not force:
        return jsonify({
            "ok": False,
            "error": "Несовместимость схемы",
            "warning": warning,
            "force_required": True,
        }), 409

    config_dir = PathLib(BASE_DIR) / "config"
    source_dir = PathLib(_source_dir_path())
    _create_version_backup("restore_before", f"Перед восстановлением backup_{backup_id}")
    if not restore_backup(PathLib(BASE_DIR), config_dir, source_dir, backup_id):
        return jsonify({"ok": False, "error": "Backup not found or restore failed"}), 404
    invalidate_metadata_cache()
    _invalidate_caches()
    get_matrix()
    get_tree()
    return jsonify({
        "ok": True,
        "message": "Данные восстановлены из бэкапа, конфиг и источник обновлены",
        "backup_id": backup_id,
    })


@app.route('/api/restore/stable', methods=["POST"])
def api_restore_stable():
    """Откат к стабильному состоянию (stable backup)."""
    if _is_db_mode():
        return jsonify({
            "ok": False,
            "error": "DB mode: use scripts/db_restore.sh",
        }), 400
    stable_id = get_stable_backup_id(PathLib(BASE_DIR))
    if not stable_id:
        return jsonify({"ok": False, "error": "Стабильное состояние не задано"}), 404
    force = bool((request.get_json() or {}).get("force"))
    # Повторяем логику api_restore для стабильного backup id
    compatible, warning = check_backup_compatibility(PathLib(BASE_DIR), stable_id)
    if not compatible and not force:
        return jsonify({
            "ok": False,
            "error": "Несовместимость схемы",
            "warning": warning,
            "force_required": True,
        }), 409
    config_dir = PathLib(BASE_DIR) / "config"
    source_dir = PathLib(_source_dir_path())
    _create_version_backup("restore_before", f"Перед восстановлением stable backup_{stable_id}")
    if not restore_backup(PathLib(BASE_DIR), config_dir, source_dir, stable_id):
        return jsonify({"ok": False, "error": "Stable backup not found or restore failed"}), 404
    invalidate_metadata_cache()
    _invalidate_caches()
    get_matrix()
    get_tree()
    return jsonify({"ok": True, "backup_id": stable_id, "message": "Восстановлено стабильное состояние"})


@app.route('/api/restore/last-change', methods=["POST"])
def api_restore_last_change():
    if _is_db_mode():
        return jsonify({
            "ok": False,
            "error": "DB mode: use approval rollback/change flow instead",
        }), 400
    """
    Откат последнего изменения.
    Так как бэкап создаётся перед мутацией, достаточно восстановить самый свежий бэкап.
    """
    backups = list_backups(PathLib(BASE_DIR))
    if not backups:
        return jsonify({"ok": False, "error": "Нет доступных бэкапов"}), 404
    last_id = backups[0]["id"]
    compatible, warning = check_backup_compatibility(PathLib(BASE_DIR), last_id)
    force = bool((request.get_json() or {}).get("force"))
    if not compatible and not force:
        return jsonify({
            "ok": False,
            "error": "Несовместимость схемы",
            "warning": warning,
            "force_required": True,
        }), 409
    config_dir = PathLib(BASE_DIR) / "config"
    source_dir = PathLib(_source_dir_path())
    _create_version_backup("restore_before", f"Перед откатом последнего изменения backup_{last_id}")
    if not restore_backup(PathLib(BASE_DIR), config_dir, source_dir, last_id):
        return jsonify({"ok": False, "error": "Backup not found or restore failed"}), 404
    invalidate_metadata_cache()
    _invalidate_caches()
    get_matrix()
    get_tree()
    return jsonify({"ok": True, "backup_id": last_id, "message": "Откат выполнен по последнему изменению"})


@app.route('/api/reload')
def api_reload():
    """Перезагрузка: сброс кэша и повторная сверка источника с чекпоинтом (при изменении файла или для autoscale)."""
    source_file = _current_source_for_backup()
    if source_file:
        _create_version_backup("reload", "Ручная перезагрузка данных")
    _invalidate_caches()
    get_matrix()
    get_tree()
    return jsonify({"ok": True, "message": "Кэш сброшен, данные загружены из источника/чекпоинта"})


@app.route('/api/autoscale/check')
def api_autoscale_check():
    """Проверка согласованности leaf-структуры matrix и tree (автоскейл)."""
    matrix = get_matrix() or {"nodes": []}
    tree = get_tree() or []
    expected = sorted(_expected_leaf_paths_from_matrix(matrix))
    actual = sorted("/".join(str(x) for x in (leaf.get("path") or [])) for leaf in collect_leaves(tree))
    expected_set = set(expected)
    actual_set = set(actual)
    missing_in_tree = sorted(expected_set - actual_set)
    extra_in_tree = sorted(actual_set - expected_set)
    return jsonify({
        "ok": len(missing_in_tree) == 0 and len(extra_in_tree) == 0,
        "expected_leaf_count": len(expected),
        "actual_leaf_count": len(actual),
        "missing_in_tree_count": len(missing_in_tree),
        "extra_in_tree_count": len(extra_in_tree),
        "missing_in_tree_sample": missing_in_tree[:20],
        "extra_in_tree_sample": extra_in_tree[:20],
    })


# ----- ОТЛАДОЧНЫЕ МАРШРУТЫ -----

@app.route('/debug')
def debug():
    cfg = get_meta()
    checkpoint = load_checkpoint(PathLib(_checkpoint_path()))
    return jsonify({
        "source_dir": cfg.get("source_dir"),
        "checkpoint_file": cfg.get("checkpoint_file"),
        "current_source": checkpoint.get("source_file") if checkpoint else None,
        "literature_dir": cfg.get("literature_dir"),
        "config": "config/settings.yaml (пути); мета — из единого источника (файл в source_dir)",
        "matrix_keys": list(get_matrix().keys()),
        "meta_keys": [k for k in cfg.keys() if k not in ("source_dir", "checkpoint_file", "literature_dir", "flexible")],
    })

# ----- СТАТИЧЕСКИЕ ФАЙЛЫ И ОБРАБОТКА ОШИБОК -----

@app.route('/static/<path:path>')
def static_files(path):
    return send_from_directory('static', path)


@app.route('/library/<path:filename>')
def library_file(filename):
    """Раздача файлов из data/library для предпросмотра."""
    if ".." in filename:
        abort(404)
    lib_dir = PathLib(_literature_dir())
    path = (lib_dir / filename).resolve()
    lib_resolved = lib_dir.resolve()
    if not str(path).startswith(str(lib_resolved)) or path == lib_resolved:
        abort(404)
    if not path.exists() or not path.is_file():
        abort(404)
    rel = path.relative_to(lib_resolved)
    return send_from_directory(lib_dir, str(rel))

@app.errorhandler(404)
def not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def server_error(e):
    return render_template('500.html'), 500

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Data Engineer Matrix')
    parser.add_argument('--port', type=int, default=5001, help='Port to run on (5000 on macOS often used by AirPlay)')
    parser.add_argument('--auto-port', action='store_true', help='Find free port automatically')
    parser.add_argument('--debug', action='store_true', help='Enable debug mode')
    args = parser.parse_args()

    os.makedirs('data', exist_ok=True)
    os.makedirs('data/sources', exist_ok=True)
    os.makedirs('data/backups', exist_ok=True)
    os.makedirs('config', exist_ok=True)
    os.makedirs('static/css', exist_ok=True)
    os.makedirs('static/js', exist_ok=True)
    os.makedirs('templates', exist_ok=True)

    # Единый источник: data/sources/matrix.json (или другой файл из source_dir). Чекпоинт — data/checkpoint.yaml
    if _is_db_mode():
        _ensure_db_schema()

    port = args.port
    if args.auto_port:
        port = find_free_port(args.port)
        if not port:
            print("❌ Не найдено свободных портов")
            sys.exit(1)

    debug_mode = args.debug or os.environ.get("DE_MATRIX_DEBUG", "0").strip().lower() in ("1", "true", "yes")
    print(f"\n🚀 Запуск на порту {port}")
    print(f"📊 Матрица: http://localhost:{port}")
    print(f"📈 Граф: http://localhost:{port}/graph")
    print(f"📋 Экспорт: http://localhost:{port}/export")
    print(f"🔧 Отладка: http://localhost:{port}/debug\n")

    app.run(debug=debug_mode, use_reloader=debug_mode, host='0.0.0.0', port=port)