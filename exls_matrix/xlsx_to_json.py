#!/usr/bin/env python3
from __future__ import annotations

import argparse
from datetime import date, datetime, time
import json
from pathlib import Path

import numpy as np

# Compatibility shim for older openpyxl versions on newer NumPy.
if not hasattr(np, "float"):
    setattr(np, "float", float)

from openpyxl import load_workbook


def _as_json_value(cell_value: object | None) -> object | None:
    if isinstance(cell_value, (datetime, date, time)):
        return cell_value.isoformat()
    return cell_value


def _to_columns(headers: tuple[object | None, ...]) -> list[str]:
    return [str(col) if col is not None else f"column_{idx + 1}" for idx, col in enumerate(headers)]


def _to_rows(values: list[tuple[object | None, ...]], columns: list[str]) -> list[dict[str, object | None]]:
    records: list[dict[str, object | None]] = []
    for row in values:
        if row is None or all(cell is None for cell in row):
            continue
        record = {column_name: _as_json_value(row[idx] if idx < len(row) else None) for idx, column_name in enumerate(columns)}
        records.append(record)
    return records


def _sheet_payload(values: list[tuple[object | None, ...]]) -> dict[str, object]:
    if not values:
        return {"columns": [], "row_count": 0, "rows": []}

    columns = _to_columns(values[0])
    rows = _to_rows(values[1:], columns)
    return {"columns": columns, "row_count": len(rows), "rows": rows}


def convert_xlsx_to_json(input_path: Path, output_path: Path) -> None:
    workbook = load_workbook(filename=input_path, data_only=False)
    result: dict[str, object] = {"source_file": input_path.name, "sheets": {}}

    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        result["sheets"][sheet.title] = _sheet_payload(values)

    output_path.write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert XLSX file to JSON.",
    )
    parser.add_argument(
        "input_xlsx",
        type=Path,
        help="Path to source XLSX file.",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Path to output JSON file. By default uses input name with .json",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path: Path = args.input_xlsx
    if not input_path.exists():
        raise FileNotFoundError(f"Input file does not exist: {input_path}")

    output_path: Path = args.output or input_path.with_suffix(".json")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    convert_xlsx_to_json(input_path, output_path)
    print(f"JSON written to: {output_path}")


if __name__ == "__main__":
    main()
